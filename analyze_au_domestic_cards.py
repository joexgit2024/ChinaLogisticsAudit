#!/usr/bin/env python3
"""
Analyze AU Domestic rate card Excel structure
"""

import pandas as pd
import openpyxl

def analyze_au_domestic_cards():
    """Analyze the structure of AU Domestic rate cards"""
    file_path = r'uploads\DHL EXPRESS AU Domestic Cards.xlsx'
    
    print("🔍 ANALYZING AU DOMESTIC RATE CARDS")
    print("=" * 60)
    
    # Load workbook
    wb = openpyxl.load_workbook(file_path)
    
    print("📋 EXCEL SHEETS:")
    for i, sheet in enumerate(wb.sheetnames, 1):
        print(f"   {i}. {sheet}")
    
    # Analyze each key sheet
    sheets_to_analyze = {
        'AU Zones TD Dom': 'City to Zone mapping',
        'AU Matrix TD Dom': 'Zone to Zone matrix', 
        'AU TD Dom': 'Rate table by weight and zone'
    }
    
    for sheet_name, description in sheets_to_analyze.items():
        if sheet_name in wb.sheetnames:
            print(f"\n🔍 SHEET: {sheet_name} ({description})")
            print("-" * 40)
            
            try:
                # Read more rows to understand structure
                df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=20, header=None)
                
                print("First 15 rows:")
                for i, row in df.iterrows():
                    if i >= 15:
                        break
                    row_str = " | ".join([str(cell) if pd.notna(cell) else "" for cell in row[:8]])
                    print(f"   {i:2d}: {row_str}")
                
                # Try to find actual data start
                if sheet_name == 'AU TD Dom':
                    print("\nLooking for rate table structure...")
                    for i, row in df.iterrows():
                        if any('Zone' in str(cell) for cell in row if pd.notna(cell)):
                            print(f"   Possible header at row {i}: {[str(cell) for cell in row if pd.notna(cell)]}")
                            break
                            
            except Exception as e:
                print(f"   Error reading sheet: {e}")
    
    # Look at specific example from the invoice
    print(f"\n🎯 EXAMPLE ANALYSIS:")
    print("From the attachment image:")
    print("   Invoice: MELR001510911")
    print("   Route: Melbourne → Sydney") 
    print("   Weight: 1.5kg")
    print("   Charged: $16.47")
    print("   Expected process:")
    print("   1. Melbourne = Zone 1 (from AU Zones TD Dom)")
    print("   2. Sydney = Zone 3 (from AU Zones TD Dom)")
    print("   3. Zone 1 → Zone 3 = Zone B (from AU Matrix TD Dom)")
    print("   4. 1.5kg in Zone B = $16.47 (from AU TD Dom)")

if __name__ == "__main__":
    analyze_au_domestic_cards()
