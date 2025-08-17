#!/usr/bin/env python3
"""
FedEx Invoice Management Routes
==============================

Enterprise-level invoice management system with filtering, sorting,
pagination, and detailed invoice views with tabbed organization.
"""

from flask import (
    Blueprint,
    request,
    render_template,
    redirect,
    url_for,
    flash,
    jsonify,
    send_file,
)
import sqlite3
import json
import pandas as pd
from datetime import datetime
import logging

# Import the FedEx unified audit engine
from fedex_unified_audit import FedExUnifiedAudit

# Import authentication
try:
    from auth_routes import require_auth, require_auth_api
except ImportError:
    def require_auth(f): return f
    def require_auth_api(f): return f

logger = logging.getLogger(__name__)

fedex_invoice_bp = Blueprint('fedex_invoices', __name__)

 
@fedex_invoice_bp.route('/fedex/invoices')
@require_auth
def invoice_list(user_data=None):
    """Main invoice list with advanced filtering and sorting"""
    
    # Get filter parameters
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 25))
    search = request.args.get('search', '').strip()
    service_filter = request.args.get('service', '').strip()
    country_filter = request.args.get('country', '').strip()
    direction_filter = request.args.get('direction', '').strip()
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()
    amount_min = request.args.get('amount_min', '').strip()
    amount_max = request.args.get('amount_max', '').strip()
    sort_by = request.args.get('sort_by', 'invoice_date')
    sort_order = request.args.get('sort_order', 'desc')
    
    conn = sqlite3.connect('fedex_audit.db')
    cursor = conn.cursor()
    
    # Build dynamic query with filters
    where_conditions = []
    params = []
    
    if search:
        where_conditions.append('''(
            invoice_no LIKE ? OR
            awb_number LIKE ? OR
            origin_country LIKE ? OR
            dest_country LIKE ? OR
            origin_loc LIKE ?
        )''')
        search_param = f'%{search}%'
        params.extend([search_param] * 5)
    
    if service_filter:
        where_conditions.append('service_type = ?')
        params.append(service_filter)
    
    if country_filter:
        where_conditions.append('(origin_country = ? OR dest_country = ?)')
        params.extend([country_filter, country_filter])
    
    if direction_filter:
        where_conditions.append('direction = ?')
        params.append(direction_filter)
    
    if date_from:
        where_conditions.append('invoice_date >= ?')
        params.append(date_from)
    
    if date_to:
        where_conditions.append('invoice_date <= ?')
        params.append(date_to)
    
    if amount_min:
        where_conditions.append('total_awb_amount_cny >= ?')
        params.append(float(amount_min))
    
    if amount_max:
        where_conditions.append('total_awb_amount_cny <= ?')
        params.append(float(amount_max))
    
    where_clause = ''
    if where_conditions:
        where_clause = 'WHERE ' + ' AND '.join(where_conditions)
    
    # Validate sort parameters
    valid_sort_columns = [
        'invoice_no', 'invoice_date', 'awb_number', 'service_type',
        'direction', 'origin_country', 'dest_country', 'pieces',
        'actual_weight_kg', 'chargeable_weight_kg', 'total_awb_amount_cny'
    ]
    
    if sort_by not in valid_sort_columns:
        sort_by = 'invoice_date'
    
    if sort_order not in ['asc', 'desc']:
        sort_order = 'desc'
    
    order_clause = f'ORDER BY {sort_by} {sort_order.upper()}'
    
    # Get total count for pagination
    count_query = f'SELECT COUNT(*) FROM fedex_invoices {where_clause}'
    cursor.execute(count_query, params)
    total_records = cursor.fetchone()[0]
    
    # Calculate pagination
    total_pages = (total_records + per_page - 1) // per_page
    offset = (page - 1) * per_page
    
    # Get invoice data
    data_query = (
        'SELECT '
        'id, invoice_no, invoice_date, awb_number, '
        'service_type, service_abbrev, direction, pieces, '
        'actual_weight_kg, chargeable_weight_kg, origin_country, '
        'dest_country, '
        'origin_loc, ship_date, delivery_datetime, rated_amount_cny, '
        'fuel_surcharge_cny, other_surcharge_cny, total_awb_amount_cny, '
        'exchange_rate '
        'FROM fedex_invoices '
        f'{where_clause} '
        f'{order_clause} '
        'LIMIT ? OFFSET ?'
    )
    
    cursor.execute(data_query, params + [per_page, offset])
    invoices = cursor.fetchall()
    
    # Get filter options for dropdowns
    cursor.execute(
        'SELECT DISTINCT service_type '
        'FROM fedex_invoices '
        'WHERE service_type IS NOT NULL '
        'ORDER BY service_type'
    )
    service_types = [row[0] for row in cursor.fetchall()]
    
    cursor.execute(
        'SELECT DISTINCT origin_country FROM fedex_invoices '
        'WHERE origin_country IS NOT NULL '
        'UNION '
        'SELECT DISTINCT dest_country FROM fedex_invoices '
        'WHERE dest_country IS NOT NULL '
        'ORDER BY 1'
    )
    countries = [row[0] for row in cursor.fetchall()]
    
    cursor.execute(
        'SELECT DISTINCT direction FROM fedex_invoices '
        'WHERE direction IS NOT NULL '
        'ORDER BY direction'
    )
    directions = [row[0] for row in cursor.fetchall()]
    
    # Get summary statistics
    stats_query = (
        'SELECT '
        'COUNT(*) as total_invoices, '
        'SUM(pieces) as total_pieces, '
        'SUM(actual_weight_kg) as total_weight, '
        'SUM(total_awb_amount_cny) as total_amount, '
        'AVG(total_awb_amount_cny) as avg_amount, '
        'MIN(invoice_date) as date_range_start, '
        'MAX(invoice_date) as date_range_end '
        'FROM fedex_invoices '
        f'{where_clause}'
    )
    
    cursor.execute(stats_query, params)
    stats = cursor.fetchone()
    
    conn.close()
    
    # Prepare template data
    template_data = {
        'invoices': invoices,
        'total_records': total_records,
        'total_pages': total_pages,
        'current_page': page,
        'per_page': per_page,
        'search': search,
        'service_filter': service_filter,
        'country_filter': country_filter,
        'direction_filter': direction_filter,
        'date_from': date_from,
        'date_to': date_to,
        'amount_min': amount_min,
        'amount_max': amount_max,
        'sort_by': sort_by,
        'sort_order': sort_order,
        'service_types': service_types,
        'countries': countries,
        'directions': directions,
        'stats': {
            'total_invoices': stats[0] or 0,
            'total_pieces': stats[1] or 0,
            'total_weight': stats[2] or 0,
            'total_amount': stats[3] or 0,
            'avg_amount': stats[4] or 0,
            'date_range_start': stats[5],
            'date_range_end': stats[6]
        }
    }
    
    return render_template('fedex_invoice_list.html', **template_data)

 
@fedex_invoice_bp.route('/fedex/invoice/<invoice_no>/<awb_number>')
@require_auth
def invoice_detail(invoice_no, awb_number, user_data=None):
    """Detailed invoice view with tabbed organization for specific AWB"""
    
    conn = sqlite3.connect('fedex_audit.db')
    cursor = conn.cursor()
    
    # Get invoice data for specific AWB
    cursor.execute(
        'SELECT * FROM fedex_invoices WHERE invoice_no = ? AND awb_number = ?',
        (invoice_no, awb_number),
    )
    row = cursor.fetchone()
    
    if not row:
        flash(f'Invoice {invoice_no} with AWB {awb_number} not found', 'error')
        return redirect(url_for('fedex_invoices.invoice_list'))
    
    # Get column names
    cursor.execute('PRAGMA table_info(fedex_invoices)')
    columns = [col[1] for col in cursor.fetchall()]
    
    # Create invoice data dictionary
    invoice_data = dict(zip(columns, row))
    
    # Get all AWBs for this invoice (for navigation)
    cursor.execute(
        'SELECT awb_number FROM fedex_invoices '
        'WHERE invoice_no = ? ORDER BY awb_number',
        (invoice_no,),
    )
    all_awbs = [row[0] for row in cursor.fetchall()]
    current_awb_index = (
        all_awbs.index(awb_number) if awb_number in all_awbs else 0
    )
    
    # Determine previous and next AWB for navigation
    prev_awb = (
        all_awbs[current_awb_index - 1] if current_awb_index > 0 else None
    )
    next_awb = (
        all_awbs[current_awb_index + 1]
        if current_awb_index < len(all_awbs) - 1
        else None
    )
    
    # Parse raw JSON if available
    raw_data = None
    if invoice_data.get('raw_json'):
        try:
            raw_data = json.loads(invoice_data['raw_json'])
        except Exception:
            raw_data = None
    
    # Get related invoices (same AWB or similar route)
    cursor.execute(
        (
            'SELECT invoice_no, invoice_date, service_type, direction, '
            'total_awb_amount_cny '
            'FROM fedex_invoices '
            'WHERE (awb_number = ? OR '
            '(origin_country = ? AND dest_country = ?)) '
            'AND invoice_no != ? '
            'ORDER BY invoice_date DESC '
            'LIMIT 10'
        ),
        (
            invoice_data['awb_number'],
            invoice_data['origin_country'],
            invoice_data['dest_country'],
            invoice_no,
        ),
    )
    
    related_invoices = cursor.fetchall()
    
    # Check for actual audit results
    has_audit_result = invoice_data.get('audit_status') is not None
    audit_status = invoice_data.get('audit_status', 'pending')
    
    # Get detailed audit information if available
    audit_results = None
    if has_audit_result:
        # Parse audit_details if it contains JSON
        audit_details_raw = invoice_data.get('audit_details')
        audit_details = None
        if audit_details_raw:
            try:
                audit_details = json.loads(audit_details_raw)
            except Exception:
                audit_details = None
        
        # Build comprehensive audit results
        audit_results = {
            'status': audit_status,
            'expected_cost_cny': invoice_data.get('expected_cost_cny'),
            'variance_cny': invoice_data.get('variance_cny'),
            'audit_timestamp': invoice_data.get('audit_timestamp'),
            'details': audit_details,
            
            # Calculate additional metrics
            'invoiced_amount': invoice_data.get('total_awb_amount_cny'),
            'variance_percentage': 0,
            'status_color': (
                'success' if audit_status == 'PASS' else (
                    'warning' if audit_status in ['OVERCHARGE', 'UNDERCHARGE']
                    else 'danger'
                )
            ),
            'status_display': (
                'REVIEW' if audit_status in ['OVERCHARGE', 'UNDERCHARGE']
                else audit_status
            )
        }
        
        # Calculate variance percentage
        if (
            audit_results['invoiced_amount']
            and audit_results['invoiced_amount'] > 0
        ):
            variance_amount = audit_results['variance_cny'] or 0
            audit_results['variance_percentage'] = (
                (variance_amount / audit_results['invoiced_amount']) * 100
            )
        
        # Add audit steps breakdown if available in details
        if audit_details and isinstance(audit_details, dict):
            audit_results['steps'] = []
            
            # Zone mapping step
            if 'zone' in audit_details:
                audit_results['steps'].append({
                    'step': 'Zone Mapping',
                    'description': (
                        'Origin: '
                        f"{invoice_data.get('origin_country')}"
                        ' → Destination: '
                        f"{invoice_data.get('dest_country')}"
                    ),
                    'result': f"Zone: {audit_details['zone']}",
                    'status': 'success'
                })
            
            # Rate lookup step
            if 'rate_per_kg' in audit_details:
                audit_results['steps'].append({
                    'step': 'Rate Lookup',
                    'description': (
                        'Service: '
                        f"{invoice_data.get('service_type')}, "
                        'Weight: '
                        f"{invoice_data.get('chargeable_weight_kg')}kg"
                    ),
                    'result': f"Rate: ¥{audit_details['rate_per_kg']}/kg",
                    'status': 'success'
                })
            
            # Base cost calculation
            if 'base_cost' in audit_details:
                audit_results['steps'].append({
                    'step': 'Base Cost Calculation',
                    'description': 'Rate × Chargeable Weight',
                    'result': f"¥{audit_details['base_cost']}",
                    'status': 'success'
                })
            
            # Surcharges step
            if 'surcharges' in audit_details:
                surcharge_total = (
                    sum(audit_details['surcharges'].values())
                    if isinstance(audit_details['surcharges'], dict)
                    else 0
                )
                audit_results['steps'].append({
                    'step': 'Surcharges Applied',
                    'description': 'Additional fees and surcharges',
                    'result': f"¥{surcharge_total}",
                    'status': 'info'
                })
            
            # Final comparison
            audit_results['steps'].append({
                'step': 'Final Comparison',
                'description': 'Expected vs Invoiced Amount',
                'result': (
                    f"Variance: ¥{audit_results['variance_cny']} ("
                    f"{audit_results['variance_percentage']:.2f}%"
                ),
                'status': audit_results['status_color']
            })
    
    # Check for invoice images (placeholder for future implementation)  
    has_invoice_image = False
    
    conn.close()
    
    return render_template('fedex_invoice_detail.html',
                         invoice_data=invoice_data,
                         invoice_no=invoice_no,
                         awb_number=awb_number,
                         all_awbs=all_awbs,
                         current_awb_index=current_awb_index + 1,  # 1-based for display
                         total_awbs=len(all_awbs),
                         prev_awb=prev_awb,
                         next_awb=next_awb,
                         raw_data=raw_data,
                         related_invoices=related_invoices,
                         has_audit_result=has_audit_result,
                         audit_results=audit_results,
                         audit_status=audit_status,
                         has_invoice_image=has_invoice_image)

# New: Detailed 8-step audit view

@fedex_invoice_bp.route('/fedex/audit-steps/<invoice_no>/<awb_number>')
@require_auth
def fedex_audit_steps(invoice_no, awb_number, user_data=None):
    """Show a transparent, step-by-step audit breakdown for a specific AWB."""
    engine = FedExUnifiedAudit()
    result = engine.audit_single_awb(invoice_no, awb_number, verbose=False)
    if not result.get('success'):
        flash(result.get('error', 'Audit failed'), 'error')
        return redirect(
            url_for(
                'fedex_invoices.invoice_detail',
                invoice_no=invoice_no,
                awb_number=awb_number,
            )
        )

    # Fetch minimal invoice fields for display and to enrich rate details
    conn = sqlite3.connect('fedex_audit.db')
    cursor = conn.cursor()
    cursor.execute(
        (
            'SELECT service_type, service_abbrev, origin_country, dest_country, '
            'total_awb_amount_cny, exchange_rate '
            'FROM fedex_invoices WHERE invoice_no = ? AND awb_number = ? LIMIT 1'
        ),
        (invoice_no, awb_number),
    )
    row = cursor.fetchone()
    conn.close()

    service_type = row[0] if row else None
    service_abbrev = row[1] if row else None
    origin_country = row[2] if row else result.get('origin_country')
    dest_country = row[3] if row else result.get('dest_country')
    claimed_cny = row[4] if row else result.get('claimed_cny')
    exchange_rate = row[5] if row else result.get('exchange_rate')

    # Try to get per-kg details if applicable
    perkg_detail = None
    try:
        eff_service = (service_abbrev or service_type or '').strip()
        rate_info = engine.get_fedex_rate(
            result['chargeable_weight_kg'],
            result['zone'],
            service_type=eff_service,
        )
        if rate_info and rate_info.get('is_per_kg'):
            perkg_detail = {
                'rate_per_kg': rate_info.get('rate_per_kg'),
                'chargeable_weight': result['chargeable_weight_kg']
            }
    except Exception:
        perkg_detail = None

    # Build the 8 steps for the UI
    steps = [
        {
            'title': '1) Zone mapping',
            'desc': f"{origin_country} → {dest_country}",
            'result': f"Zone {result['zone']}"
        },
        {
            'title': '2) Weight determination',
            'desc': (
                'Use chargeable weight when available; '
                'apply FedEx rounding rules'
            ),
            'result': (
                f"Chargeable: {result['chargeable_weight_kg']} kg (Actual: "
                f"{result['actual_weight_kg']} kg)"
            ),
        },
        {
            'title': '3) Base rate lookup',
            'desc': f"Service: {service_abbrev or service_type}",
            'result': (
                (
                    f"${perkg_detail['rate_per_kg']:.2f}/kg × "
                    f"{perkg_detail['chargeable_weight']}kg = "
                    f"${result['base_cost_usd']:.2f}"
                )
                if perkg_detail and perkg_detail.get('rate_per_kg') is not None
                else (
                    f"${result['base_cost_usd']:.2f} "
                    f"(rate type: {result['rate_type']})"
                )
            ),
        },
        {
            'title': '4) Fuel surcharge',
            'desc': 'Applied at 25.5% of base',
            'result': f"${result['fuel_surcharge_usd']:.2f}"
        },
        {
            'title': '5) Subtotal (USD)',
            'desc': 'Base + Fuel surcharge',
            'result': f"${result['subtotal_usd']:.2f}"
        },
        {
            'title': '6) Currency conversion',
            'desc': f"USD → CNY at FX {exchange_rate}",
            'result': f"¥{result['subtotal_cny']:.2f}"
        },
        {
            'title': '7) VAT (6%)',
            'desc': 'VAT applied to CNY subtotal',
            'result': f"¥{result['vat_cny']:.2f}"
        },
        {
            'title': '8) Final total & variance',
            'desc': 'Compare expected vs invoiced',
            'result': (
                f"Expected: ¥{result['total_expected_cny']:.2f} | "
                f"Claimed: ¥{claimed_cny:.2f} | "
                f"Variance: ¥{result['variance_cny']:.2f} "
                f"({result['variance_percent']:+.1f}%) — "
                f"{result['audit_status']}"
            ),
        },
    ]

    return render_template(
        'fedex_audit_steps.html',
        invoice_no=invoice_no,
        awb_number=awb_number,
        service_type=service_type,
        service_abbrev=service_abbrev,
        origin_country=origin_country,
        dest_country=dest_country,
        claimed_cny=claimed_cny,
        exchange_rate=exchange_rate,
        result=result,
        steps=steps
    )

@fedex_invoice_bp.route('/fedex/invoice/<invoice_no>')
@require_auth
def invoice_detail_redirect(invoice_no, user_data=None):
    """Redirect old invoice URLs to first AWB for backward compatibility"""
    
    conn = sqlite3.connect('fedex_audit.db')
    cursor = conn.cursor()
    
    # Get first AWB for this invoice
    cursor.execute(
        (
            'SELECT awb_number FROM fedex_invoices '
            'WHERE invoice_no = ? ORDER BY awb_number LIMIT 1'
        ),
        (invoice_no,),
    )
    row = cursor.fetchone()
    conn.close()
    
    if not row:
        flash(f'Invoice {invoice_no} not found', 'error')
        return redirect(url_for('fedex_invoices.invoice_list'))
    
    # Redirect to specific AWB URL
    return redirect(
        url_for(
            'fedex_invoices.invoice_detail',
            invoice_no=invoice_no,
            awb_number=row[0],
        )
    )

@fedex_invoice_bp.route('/fedex/invoices/export')
@require_auth
def export_invoices(user_data=None):
    """Export filtered invoice data to Excel with auto-fit columns"""
    # Get same filters as invoice list
    search = request.args.get('search', '').strip()
    service_filter = request.args.get('service_type', '')
    country_filter = request.args.get('country', '')
    direction_filter = request.args.get('direction', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    amount_min = request.args.get('amount_min', '')
    amount_max = request.args.get('amount_max', '')
    
    try:
        conn = sqlite3.connect('fedex_audit.db')
        cursor = conn.cursor()
        
        # Build same query as invoice list but get all results
        where_conditions = []
        params = []
        
        if search:
            where_conditions.append('''(
                invoice_no LIKE ? OR 
                awb_number LIKE ? OR 
                origin_country LIKE ? OR
                dest_country LIKE ? OR
                origin_loc LIKE ?
            )''')
            search_param = f'%{search}%'
            params.extend([search_param] * 5)
        
        if service_filter:
            where_conditions.append('service_type = ?')
            params.append(service_filter)
        
        if country_filter:
            where_conditions.append('(origin_country = ? OR dest_country = ?)')
            params.extend([country_filter, country_filter])
        
        if direction_filter:
            where_conditions.append('direction = ?')
            params.append(direction_filter)
        
        if date_from:
            where_conditions.append('invoice_date >= ?')
            params.append(date_from)
        
        if date_to:
            where_conditions.append('invoice_date <= ?')
            params.append(date_to)
        
        if amount_min:
            where_conditions.append('total_awb_amount_cny >= ?')
            params.append(float(amount_min))
        
        if amount_max:
            where_conditions.append('total_awb_amount_cny <= ?')
            params.append(float(amount_max))
        
        where_clause = ''
        if where_conditions:
            where_clause = 'WHERE ' + ' AND '.join(where_conditions)
        
        # Get all invoices for export
        query = f'''
            SELECT invoice_no, invoice_date, awb_number, service_type, service_abbrev,
                   direction, pieces, actual_weight_kg, chargeable_weight_kg, dim_weight_kg,
                   origin_country, dest_country, origin_loc, ship_date, delivery_datetime,
                   exchange_rate, rated_amount_cny, discount_amount_cny, fuel_surcharge_cny,
                   other_surcharge_cny, vat_amount_cny, total_awb_amount_cny,
                   audit_status, expected_cost_cny, variance_cny, audit_timestamp
            FROM fedex_invoices 
            {where_clause}
            ORDER BY invoice_date DESC, invoice_no, awb_number
        '''
        
        cursor.execute(query, params)
        results = cursor.fetchall()
        conn.close()
        
        if not results:
            flash('No invoices found to export', 'warning')
            return redirect(url_for('fedex_invoices.invoice_list'))
        
        # Create DataFrame
        df = pd.DataFrame(results, columns=[
            'Invoice No', 'Invoice Date', 'AWB Number', 'Service Type', 'Service Abbrev',
            'Direction', 'Pieces', 'Actual Weight (kg)', 'Chargeable Weight (kg)', 'Dim Weight (kg)',
            'Origin Country', 'Dest Country', 'Origin Location', 'Ship Date', 'Delivery Date',
            'Exchange Rate', 'Rated Amount (CNY)', 'Discount Amount (CNY)', 'Fuel Surcharge (CNY)',
            'Other Surcharge (CNY)', 'VAT Amount (CNY)', 'Total Amount (CNY)',
            'Audit Status', 'Expected Cost (CNY)', 'Variance (CNY)', 'Audit Timestamp'
        ])
        
        # Create Excel file with auto-fit columns
        from io import BytesIO
        from openpyxl.styles import Font, PatternFill, Alignment
        
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='FedEx Invoices', index=False)
            
            # Auto-fit columns with professional styling
            def auto_fit_columns(worksheet):
                # Style the header row
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    # Style header cell
                    header_cell = worksheet[f"{column_letter}1"]
                    header_cell.font = header_font
                    header_cell.fill = header_fill
                    header_cell.alignment = header_alignment
                    
                    # Calculate column width
                    for cell in column:
                        try:
                            cell_length = len(str(cell.value)) if cell.value is not None else 0
                            if cell_length > max_length:
                                max_length = cell_length
                        except:
                            pass
                    
                    # Set column width with padding (minimum 12, maximum 60)
                    adjusted_width = min(max(max_length + 3, 12), 60)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Apply formatting to the sheet
            auto_fit_columns(writer.sheets['FedEx Invoices'])
        
        # Prepare download
        buffer.seek(0)
        filename = f"fedex_invoices_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'Export failed: {str(e)}', 'error')
        return redirect(url_for('fedex_invoices.invoice_list'))

@fedex_invoice_bp.route('/fedex/api/invoice-summary/<invoice_no>')
@require_auth_api
def invoice_summary_api(invoice_no, user_data=None):
    """API endpoint for quick invoice summary"""
    
    conn = sqlite3.connect('fedex_audit.db')
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT invoice_no, invoice_date, service_type, direction,
               origin_country, dest_country, pieces, actual_weight_kg,
               total_awb_amount_cny
        FROM fedex_invoices 
        WHERE invoice_no = ?
    ''', (invoice_no,))
    
    row = cursor.fetchone()
    conn.close()
    
    if not row:
        return jsonify({'error': 'Invoice not found'}), 404
    
    return jsonify({
        'invoice_no': row[0],
        'invoice_date': row[1],
        'service_type': row[2],
        'direction': row[3],
        'origin_country': row[4],
        'dest_country': row[5],
        'pieces': row[6],
        'weight_kg': row[7],
        'total_amount_cny': row[8]
    })

@fedex_invoice_bp.route('/fedex/api/invoices/stats')
@require_auth_api
def invoices_stats_api(user_data=None):
    """API endpoint for invoice statistics"""
    
    conn = sqlite3.connect('fedex_audit.db')
    cursor = conn.cursor()
    
    # Get overall statistics
    cursor.execute('''
        SELECT 
            COUNT(*) as total_invoices,
            SUM(pieces) as total_pieces,
            SUM(actual_weight_kg) as total_weight,
            SUM(total_awb_amount_cny) as total_amount,
            COUNT(DISTINCT service_type) as service_types,
            COUNT(DISTINCT origin_country) as origin_countries,
            COUNT(DISTINCT dest_country) as dest_countries
        FROM fedex_invoices
    ''')
    
    stats = cursor.fetchone()
    
    # Get service breakdown
    cursor.execute('''
        SELECT service_type, COUNT(*), SUM(total_awb_amount_cny)
        FROM fedex_invoices 
        WHERE service_type IS NOT NULL
        GROUP BY service_type
        ORDER BY COUNT(*) DESC
    ''')
    
    service_breakdown = cursor.fetchall()
    
    # Get monthly trends
    cursor.execute('''
        SELECT 
            strftime('%Y-%m', invoice_date) as month,
            COUNT(*) as invoice_count,
            SUM(total_awb_amount_cny) as total_amount
        FROM fedex_invoices 
        WHERE invoice_date IS NOT NULL
        GROUP BY strftime('%Y-%m', invoice_date)
        ORDER BY month DESC
        LIMIT 12
    ''')
    
    monthly_trends = cursor.fetchall()
    
    conn.close()
    
    return jsonify({
        'overall_stats': {
            'total_invoices': stats[0] or 0,
            'total_pieces': stats[1] or 0,
            'total_weight': stats[2] or 0,
            'total_amount': stats[3] or 0,
            'service_types': stats[4] or 0,
            'origin_countries': stats[5] or 0,
            'dest_countries': stats[6] or 0
        },
        'service_breakdown': [
            {'service': row[0], 'count': row[1], 'amount': row[2]}
            for row in service_breakdown
        ],
        'monthly_trends': [
            {'month': row[0], 'count': row[1], 'amount': row[2]}
            for row in monthly_trends
        ]
    })


# =====================================================
# FedEx Batch Audit Routes
# =====================================================

@fedex_invoice_bp.route('/fedex/batch-audit')
# @require_auth  # Temporarily disabled for testing
def batch_audit_dashboard(user_data=None):
    """Display FedEx batch audit dashboard"""
    
    # Get audit status summary directly from database
    conn = sqlite3.connect('fedex_audit.db')
    cursor = conn.cursor()
    
    # Total invoices and AWBs
    cursor.execute('SELECT COUNT(DISTINCT invoice_no) FROM fedex_invoices')
    total_invoices = cursor.fetchone()[0]
    
    cursor.execute('SELECT COUNT(*) FROM fedex_invoices')
    total_awbs = cursor.fetchone()[0]
    
    # Audited AWBs
    cursor.execute('SELECT COUNT(*) FROM fedex_invoices WHERE audit_status IS NOT NULL')
    audited_awbs = cursor.fetchone()[0]
    
    # Unaudited AWBs
    unaudited_awbs = total_awbs - audited_awbs
    
    # Audit completion rate
    audit_completion_rate = (audited_awbs / total_awbs * 100) if total_awbs > 0 else 0
    
    # Total amount
    cursor.execute('SELECT SUM(total_awb_amount_cny) FROM fedex_invoices')
    total_amount = cursor.fetchone()[0] or 0
    
    # Audit status counts
    cursor.execute('''
        SELECT audit_status, COUNT(*) 
        FROM fedex_invoices 
        WHERE audit_status IS NOT NULL 
        GROUP BY audit_status
    ''')
    status_counts = dict(cursor.fetchall())
    
    # Total variance calculation
    cursor.execute('SELECT SUM(variance_cny) FROM fedex_invoices WHERE variance_cny IS NOT NULL')
    total_variance = cursor.fetchone()[0] or 0
    
    # Get some unaudited invoices for display
    cursor.execute('''
        SELECT invoice_no, awb_number, origin_country, dest_country, total_awb_amount_cny
        FROM fedex_invoices 
        WHERE audit_status IS NULL 
        ORDER BY invoice_no, awb_number
        LIMIT 10
    ''')
    unaudited_invoices = []
    for row in cursor.fetchall():
        unaudited_invoices.append({
            'invoice_no': row[0],
            'awb_number': row[1],
            'origin_country': row[2],
            'dest_country': row[3],
            'route': f"{row[2]} - {row[3]}",
            'amount': row[4] or 0
        })
    
    conn.close()
    
    # Build summary object with proper mapping
    pass_count = status_counts.get('PASS', 0)
    overcharge_count = status_counts.get('OVERCHARGE', 0)
    undercharge_count = status_counts.get('UNDERCHARGE', 0)
    fail_count = status_counts.get('FAIL', 0)
    
    # Map OVERCHARGE and UNDERCHARGE to REVIEW
    review_count = overcharge_count + undercharge_count
    
    audit_summary = {
        'total_invoices': total_invoices,
        'audited_awbs': audited_awbs,
        'unaudited_awbs': unaudited_awbs,
        'audit_completion_rate': audit_completion_rate,
        'total_amount': total_amount,
        'pass_count': pass_count,
        'review_count': review_count,  # OVERCHARGE + UNDERCHARGE
        'fail_count': fail_count,
        'total_variance': total_variance,
        # Keep individual counts for debugging
        'overcharge_count': overcharge_count,
        'undercharge_count': undercharge_count
    }
    
    return render_template('fedex_batch_audit.html', 
                         summary=audit_summary,
                         unaudited_count=len(unaudited_invoices),
                         unaudited_invoices=unaudited_invoices)

@fedex_invoice_bp.route('/fedex/batch-audit/status')
def get_batch_audit_status():
    """Get current FedEx batch audit status (API endpoint)"""
    conn = sqlite3.connect('fedex_audit.db')
    cursor = conn.cursor()
    
    # Get basic counts
    cursor.execute('SELECT COUNT(DISTINCT invoice_no) FROM fedex_invoices')
    total_invoices = cursor.fetchone()[0]
    
    cursor.execute('SELECT COUNT(*) FROM fedex_invoices WHERE audit_status IS NOT NULL')
    audited_awbs = cursor.fetchone()[0]
    
    cursor.execute('SELECT COUNT(*) FROM fedex_invoices WHERE audit_status IS NULL')
    unaudited_awbs = cursor.fetchone()[0]
    
    # Get status counts
    cursor.execute('SELECT audit_status, COUNT(*) FROM fedex_invoices WHERE audit_status IS NOT NULL GROUP BY audit_status')
    status_counts = dict(cursor.fetchall())
    
    # Get some unaudited invoices
    cursor.execute('SELECT invoice_no, awb_number FROM fedex_invoices WHERE audit_status IS NULL LIMIT 5')
    unaudited_invoices = [{'invoice_no': row[0], 'awb_number': row[1]} for row in cursor.fetchall()]
    
    conn.close()
    
    summary = {
        'total_invoices': total_invoices,
        'audited_awbs': audited_awbs,
        'unaudited_awbs': unaudited_awbs,
        'pass_count': status_counts.get('PASS', 0),
        'review_count': status_counts.get('OVERCHARGE', 0) + status_counts.get('UNDERCHARGE', 0),
        'fail_count': status_counts.get('FAIL', 0)
    }
    
    return jsonify({
        'success': True,
        'summary': summary,
        'unaudited_invoices': len(unaudited_invoices),
        'sample_unaudited': unaudited_invoices[:5]
    })

@fedex_invoice_bp.route('/fedex/batch-audit/run', methods=['POST'])
def run_batch_audit():
    """Run FedEx batch audit on all unaudited invoices"""
    try:
        engine = FedExUnifiedAudit()
        
        # Get request parameters
        data = request.get_json() or {}
        audit_type = data.get('audit_type', 'all_unaudited')
        specific_invoices = data.get('invoice_list', [])
        
        if audit_type == 'specific_invoices' and specific_invoices:
            # Accept both list of invoice numbers or objects
            invoice_nos = []
            for item in specific_invoices:
                if isinstance(item, dict):
                    inv = item.get('invoice_no')
                    if inv and inv not in invoice_nos:
                        invoice_nos.append(inv)
                else:
                    invoice_nos.append(item)
            if not invoice_nos:
                return jsonify({'success': False, 'error': 'No invoice numbers provided'}), 400
            # Audit at invoice level
            result = engine.audit_batch_invoices(invoice_nos)
        else:
            # Audit all unaudited invoices  
            conn = sqlite3.connect('fedex_audit.db')
            cursor = conn.cursor()
            cursor.execute('SELECT DISTINCT invoice_no FROM fedex_invoices WHERE audit_status IS NULL')
            unaudited_invoices = [row[0] for row in cursor.fetchall()]
            conn.close()
            
            result = engine.audit_batch_invoices(unaudited_invoices)
        
        return jsonify({
            'success': True,
            'batch_result': result
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@fedex_invoice_bp.route('/fedex/batch-audit/audit-awb', methods=['POST'])
def audit_single_awb_api():
    """Audit a single AWB and update its record, returning detailed debug info."""
    try:
        data = request.get_json() or {}
        invoice_no = str(data.get('invoice_no', '')).strip()
        awb_number = str(data.get('awb_number', '')).strip()
        if not invoice_no or not awb_number:
            return jsonify({'success': False, 'error': 'invoice_no and awb_number are required'}), 400

        auditor = FedExUnifiedAudit()
        result = auditor.audit_single_awb(invoice_no, awb_number, verbose=True)

        if not result.get('success'):
            logger.warning("Single AWB audit failed: %s", result)
            return jsonify({'success': False, 'error': result.get('error', 'Unknown error'), 'details': result}), 200

        # Update this AWB in DB
        conn = sqlite3.connect('fedex_audit.db')
        cursor = conn.cursor()
        cursor.execute('''
            UPDATE fedex_invoices 
            SET audit_status = ?,
                expected_cost_cny = ?,
                variance_cny = ?,
                audit_timestamp = ?,
                audit_details = ?
            WHERE invoice_no = ? AND awb_number = ?
        ''', (
            result['audit_status'],
            result['total_expected_cny'],
            result['variance_cny'],
            datetime.now().isoformat(),
            f"Zone {result['zone']}, {result['chargeable_weight_kg']}kg, Rate: ${result['base_cost_usd']:.2f}",
            invoice_no,
            awb_number
        ))
        conn.commit()
        conn.close()

        return jsonify({'success': True, 'audit': result})
    except Exception as e:
        logger.exception('Error auditing single AWB')
        return jsonify({'success': False, 'error': str(e)}), 500

@fedex_invoice_bp.route('/fedex/batch-audit/rerun', methods=['POST'])
def rerun_batch_audit():
    """Re-run FedEx batch audit on specific invoices"""
    try:
        engine = FedExUnifiedAudit()
        
        data = request.get_json() or {}
        invoice_list = data.get('invoice_list', [])
        
        if not invoice_list:
            # If no specific invoices provided, get all invoices
            conn = sqlite3.connect('fedex_audit.db')
            cursor = conn.cursor()
            cursor.execute('SELECT DISTINCT invoice_no FROM fedex_invoices')
            invoice_list = [row[0] for row in cursor.fetchall()]
            conn.close()
        
        result = engine.audit_batch_invoices(invoice_list)
        
        return jsonify({
            'success': True,
            'message': f'Successfully re-audited {len(invoice_list)} invoices',
            'processed_count': len(invoice_list),
            'result': result
        })
        
    except Exception as e:
        logger.error(f"Error in rerun_batch_audit: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@fedex_invoice_bp.route('/fedex/batch-audit/rerun-all', methods=['POST'])
# @require_auth_api  # Temporarily disabled for testing
def rerun_all_audits(user_data=None):
    """Re-run FedEx batch audit on ALL invoices using the unified audit system"""
    try:
        logger.info("Starting rerun_all_audits with unified audit system")
        
        # Import our proven unified audit system
        from fedex_unified_audit import FedExUnifiedAudit
        auditor = FedExUnifiedAudit()
        
        # Get ALL unique invoices from the database
        conn = sqlite3.connect('fedex_audit.db')
        cursor = conn.cursor()
        logger.info("Connected to database")
        
        # Clear existing audit results
        cursor.execute('''
            UPDATE fedex_invoices 
            SET audit_status = NULL, 
                expected_cost_cny = NULL, 
                variance_cny = NULL, 
                audit_timestamp = NULL, 
                audit_details = NULL
        ''')
        conn.commit()
        logger.info("Cleared existing audit results")
        
        # Get all unique invoices
        cursor.execute('SELECT DISTINCT invoice_no FROM fedex_invoices ORDER BY invoice_no')
        invoice_numbers = [row[0] for row in cursor.fetchall()]
        conn.close()
        
        if not invoice_numbers:
            return jsonify({
                'success': False,
                'error': 'No invoices found to audit'
            }), 400
        
        logger.info(f"Found {len(invoice_numbers)} invoices to audit")
        
        # Audit each invoice using our proven unified system
        success_count = 0
        error_count = 0
        total_variance = 0
        
        for invoice_no in invoice_numbers:
            try:
                # Audit this invoice
                result = auditor.audit_invoice(invoice_no, verbose=False)
                
                if result['success']:
                    # Update database with audit results
                    if auditor.update_audit_results(result):
                        success_count += 1
                        total_variance += result['total_variance_cny']
                        logger.info(f"Successfully audited invoice {invoice_no} with {result['awb_count']} AWBs")
                    else:
                        error_count += 1
                        logger.error(f"Failed to save audit results for invoice {invoice_no}")
                else:
                    error_count += 1
                    logger.error(f"Failed to audit invoice {invoice_no}: {result.get('error', 'Unknown error')}")
                    
            except Exception as e:
                error_count += 1
                logger.error(f"Exception auditing invoice {invoice_no}: {str(e)}")
        
        # Get final audit statistics
        conn = sqlite3.connect('fedex_audit.db')
        cursor = conn.cursor()
        
        cursor.execute('SELECT COUNT(*) FROM fedex_invoices WHERE audit_status IS NOT NULL')
        total_audited = cursor.fetchone()[0]
        
        cursor.execute('''
            SELECT audit_status, COUNT(*) 
            FROM fedex_invoices 
            WHERE audit_status IS NOT NULL 
            GROUP BY audit_status
        ''')
        status_breakdown = dict(cursor.fetchall())
        
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'Batch audit completed. {success_count} invoices successful, {error_count} errors.',
            'total_processed': len(invoice_numbers),
            'successful_invoices': success_count,
            'failed_invoices': error_count,
            'total_audited_awbs': total_audited,
            'total_variance_cny': round(total_variance, 2),
            'status_breakdown': status_breakdown
        })
        
    except Exception as e:
        logger.error(f"Error in rerun_all_audits: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@fedex_invoice_bp.route('/fedex/batch-audit/results')
# @require_auth  # Temporarily disabled for testing
def batch_audit_results(user_data=None):
    """Display FedEx batch audit results with filtering and pagination"""
    # Get query parameters
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 25, type=int)
    status_filter = request.args.get('status', 'all')
    sort_by = request.args.get('sort_by', 'audit_timestamp')
    sort_order = request.args.get('sort_order', 'desc')
    datatables_mode = request.args.get('datatables', 'true').lower() == 'true'  # Enable DataTables by default
    
    conn = sqlite3.connect('fedex_audit.db')
    cursor = conn.cursor()
    
    # Build WHERE clause for status filter
    where_clause = "WHERE audit_status IS NOT NULL"
    params = []
    
    if status_filter != 'all':
        # Map filter values to database values
        # Note: UNDERCHARGE should be treated as PASS in the UI
        if status_filter.upper() == 'REVIEW':
            # Only OVERCHARGE is considered REVIEW in UI
            where_clause += " AND audit_status = ?"
            params.append('OVERCHARGE')
        elif status_filter.upper() == 'PASS':
            # Include both PASS and UNDERCHARGE in PASS bucket
            where_clause += " AND audit_status IN (?, ?)"
            params.extend(['PASS', 'UNDERCHARGE'])
        elif status_filter.upper() == 'FAIL':
            where_clause += " AND audit_status = ?"
            params.append('FAIL')
        else:
            # Fallback for exact match (case-insensitive)
            where_clause += " AND UPPER(audit_status) = UPPER(?)"
            params.append(status_filter)
    
    # Build ORDER BY clause
    order_clause = f"ORDER BY {sort_by} {sort_order.upper()}"
    
    # Get total count
    count_query = f"""
        SELECT COUNT(*) 
        FROM fedex_invoices 
        {where_clause}
    """
    cursor.execute(count_query, params)
    total_count = cursor.fetchone()[0]
    
    # Calculate offset and pagination
    if datatables_mode:
        # In DataTables mode, return all results
        # (DataTables handles pagination client-side)
        offset = 0
        limit_clause = ""
        per_page = total_count  # Set to total count for pagination display
    else:
        # Traditional server-side pagination
        offset = (page - 1) * per_page
        limit_clause = "LIMIT ? OFFSET ?"
    
    total_pages = (
        max(1, (total_count + per_page - 1) // per_page)
        if per_page > 0 else 1
    )
    
    # Get audit results
    results_query = (
        'SELECT invoice_no, awb_number, service_type, origin_country, '
        'dest_country, actual_weight_kg, chargeable_weight_kg, '
        'total_awb_amount_cny, expected_cost_cny, variance_cny, '
        'audit_status, audit_timestamp, audit_details '
        'FROM fedex_invoices '
        f'{where_clause} '
        f'{order_clause} '
        f'{limit_clause}'
    )
    
    if datatables_mode:
        # Return all results for DataTables
        cursor.execute(results_query, params)
    else:
        # Traditional pagination
        cursor.execute(results_query, params + [per_page, offset])
    
    results = []
    for row in cursor.fetchall():
        # Calculate route string
        route = f"{row[3]} → {row[4]}" if row[3] and row[4] else "-"
        
        # Calculate variance percentage
        variance_percentage = 0
        if row[7] and row[7] > 0:  # total_awb_amount_cny (now at index 7)
            variance_percentage = (row[9] / row[7] * 100) if row[9] else 0
        
        results.append({
            'invoice_no': row[0],
            'awb_number': row[1],
            'service_type': row[2],
            'origin_country': row[3],
            'dest_country': row[4],
            'actual_weight_kg': row[5],
            'chargeable_weight_kg': row[6],
            'invoiced_amount': row[7],  # total_awb_amount_cny
            'expected_amount': row[8],  # expected_cost_cny
            'variance': row[9],         # Map variance_cny to variance
            'audit_status': row[10],
            'audit_timestamp': row[11],
            'audit_details': row[12],
            
            # Template-specific mappings
            # Treat UNDERCHARGE as PASS for UI status
            'status': (
                'PASS' if row[10] in ['PASS', 'UNDERCHARGE']
                else ('REVIEW' if row[10] == 'OVERCHARGE' else row[10])
            ),
            'detailed_status': row[10],  # Original status for color
            'route': route,
            'actual_weight': row[5],
            'chargeable_weight': row[6],  # Use chargeable weight
            'variance_amount': row[9],   # Template expects 'variance_amount'
            'variance_percentage': variance_percentage,
            
            # Keep original field names for compatibility
            'total_awb_amount_cny': row[7],
            'expected_cost_cny': row[8],
            'variance_cny': row[9]
        })
    
    # Get summary statistics
    cursor.execute(
        'SELECT audit_status, COUNT(*), SUM(total_awb_amount_cny), '
        'SUM(variance_cny) FROM fedex_invoices '
        'WHERE audit_status IS NOT NULL '
        'GROUP BY audit_status'
    )
    status_data = cursor.fetchall()
    
    # Initialize counters
    pass_count = 0
    review_count = 0
    fail_count = 0
    total_amount = 0
    total_variance = 0
    
    for status, count, amount, variance in status_data:
        if status in ['PASS', 'UNDERCHARGE']:
            # UNDERCHARGE contributes to PASS bucket
            pass_count += count
            total_amount += amount or 0
            total_variance += variance or 0
        elif status == 'OVERCHARGE':
            # Only OVERCHARGE is REVIEW
            review_count += count
            total_amount += amount or 0
            total_variance += variance or 0
        elif status == 'FAIL':
            fail_count += count
            total_amount += amount or 0
            total_variance += variance or 0
    
    summary = {
        'pass_count': pass_count,
        'review_count': review_count,
        'fail_count': fail_count,
        'total_amount': total_amount,
        'total_variance': total_variance
    }
    
    conn.close()
    
    pagination_info = {
        'page': page,
        'per_page': per_page,
        'total': total_count,
        'total_pages': total_pages,
        'has_prev': page > 1,
        'has_next': page < total_pages,
        'prev_num': page - 1 if page > 1 else None,
        'next_num': page + 1 if page < total_pages else None
    }
    
    return render_template('fedex_batch_audit_results.html',
                         results=results,
                         pagination=pagination_info,
                         summary=summary,
                         datatables_mode=datatables_mode,
                         current_filter={'status': status_filter, 'sort_by': sort_by, 'sort_order': sort_order})

@fedex_invoice_bp.route('/fedex/batch-audit/export')
def export_audit_results():
    """Export FedEx audit results to Excel"""
    try:
        from io import BytesIO
        import pandas as pd
        from flask import send_file, make_response
        
        # Get query parameters
        status_filter = request.args.get('status', 'all')
        
        conn = sqlite3.connect('fedex_audit.db')
        cursor = conn.cursor()
        
        # Build WHERE clause for status filter
        where_clause = "WHERE audit_status IS NOT NULL"
        params = []
        
        if status_filter != 'all':
            # Map filter values to database values (same as results view)
            if status_filter.upper() == 'REVIEW':
                where_clause += " AND audit_status IN (?, ?)"
                params.extend(['OVERCHARGE', 'UNDERCHARGE'])
            elif status_filter.upper() == 'PASS':
                where_clause += " AND audit_status = ?"
                params.append('PASS')
            elif status_filter.upper() == 'FAIL':
                where_clause += " AND audit_status = ?"
                params.append('FAIL')
            else:
                # Fallback for exact match (case-insensitive)
                where_clause += " AND UPPER(audit_status) = UPPER(?)"
                params.append(status_filter)
        
        # Get all audit results for export
        query = f"""
            SELECT invoice_no, awb_number, service_type, origin_country, dest_country, 
                   actual_weight_kg, total_awb_amount_cny, expected_cost_cny, variance_cny, 
                   audit_status, audit_timestamp, audit_details
            FROM fedex_invoices 
            {where_clause}
            ORDER BY audit_timestamp DESC
        """
        cursor.execute(query, params)
        results = cursor.fetchall()
        conn.close()
        
        if not results:
            flash('No audit results found to export', 'warning')
            return redirect(url_for('fedex_invoices.batch_audit_results'))
        
        # Create DataFrames for export
        results_df = pd.DataFrame([
            {
                'Invoice No': row[0],
                'AWB Number': row[1], 
                'Service Type': row[2],
                'Origin': row[3],
                'Destination': row[4],
                'Weight (kg)': row[5],
                'Invoiced Amount (CNY)': row[6],
                'Expected Amount (CNY)': row[7],
                'Variance (CNY)': row[8],
                'Variance %': f"{(row[8] / row[6] * 100) if row[6] else 0:.2f}%",
                'Status': row[9],
                'Audit Date': row[10],
                'Audit Details': row[11]
            }
            for row in results
        ])
        
        # Summary statistics
        conn = sqlite3.connect('fedex_audit.db')
        cursor = conn.cursor()
        
        cursor.execute('SELECT COUNT(DISTINCT invoice_no) FROM fedex_invoices')
        total_invoices = cursor.fetchone()[0]
        
        cursor.execute('SELECT COUNT(*) FROM fedex_invoices WHERE audit_status IS NOT NULL')
        audited_awbs = cursor.fetchone()[0]
        
        cursor.execute('SELECT SUM(total_awb_amount_cny), SUM(expected_cost_cny), SUM(variance_cny) FROM fedex_invoices WHERE audit_status IS NOT NULL')
        amounts = cursor.fetchone()
        total_amount = amounts[0] or 0
        total_expected = amounts[1] or 0
        total_variance = amounts[2] or 0
        
        cursor.execute('SELECT audit_status, COUNT(*) FROM fedex_invoices WHERE audit_status IS NOT NULL GROUP BY audit_status')
        status_counts = dict(cursor.fetchall())
        
        conn.close()
        
        # Summary dataframe
        summary_df = pd.DataFrame([
            {'Metric': 'Total Invoices', 'Value': total_invoices},
            {'Metric': 'Audited AWBs', 'Value': audited_awbs},
            {'Metric': 'Pass Count', 'Value': status_counts.get('PASS', 0)},
            {'Metric': 'Overcharge Count', 'Value': status_counts.get('OVERCHARGE', 0)},
            {'Metric': 'Undercharge Count', 'Value': status_counts.get('UNDERCHARGE', 0)},
            {'Metric': 'Fail Count', 'Value': status_counts.get('FAIL', 0)},
            {'Metric': 'Total Amount (CNY)', 'Value': f"¥{total_amount:,.2f}"},
            {'Metric': 'Total Expected (CNY)', 'Value': f"¥{total_expected:,.2f}"},
            {'Metric': 'Total Variance (CNY)', 'Value': f"¥{total_variance:,.2f}"}
        ])
        
        # Create Excel file
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            results_df.to_excel(writer, sheet_name='Audit Results', index=False)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Auto-fit columns for all sheets
            def auto_fit_columns(worksheet):
                from openpyxl.styles import Font, PatternFill, Alignment
                
                # Style the header row
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    # Style header cell
                    header_cell = worksheet[f"{column_letter}1"]
                    header_cell.font = header_font
                    header_cell.fill = header_fill
                    header_cell.alignment = header_alignment
                    
                    # Calculate column width
                    for cell in column:
                        try:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                        except:
                            pass
                    adjusted_width = min(max_length + 3, 60)  # Add padding and max width
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Apply auto-fit to all sheets
            for sheet_name in writer.sheets:
                auto_fit_columns(writer.sheets[sheet_name])
        
        buffer.seek(0)
        
        # Create response
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'FedEx_Audit_Results_{timestamp}.xlsx'
        
        response = make_response(send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        ))
        
        return response
        
    except Exception as e:
        flash(f'Error exporting audit results: {str(e)}', 'error')
    return redirect(url_for('fedex_invoices.batch_audit_results'))

@fedex_invoice_bp.route('/fedex/batch-audit/export-detailed')
def export_detailed_audit_results():
    """Export detailed FedEx audit results with breakdown"""
    try:
        from io import BytesIO
        import pandas as pd
        from flask import send_file, make_response
        
        # Get query parameters
        status_filter = request.args.get('status', 'all')
        
        conn = sqlite3.connect('fedex_audit.db')
        cursor = conn.cursor()
        
        # Build WHERE clause for status filter
        where_clause = "WHERE audit_status IS NOT NULL"
        params = []
        
        if status_filter != 'all':
            where_clause += " AND audit_status = ?"
            params.append(status_filter)
        
        # Get all audit results for detailed export with additional columns
        query = f"""
            SELECT invoice_no, awb_number, service_type, origin_country, dest_country, 
                   actual_weight_kg, total_awb_amount_cny, expected_cost_cny, variance_cny, 
                   audit_status, audit_timestamp, exchange_rate, fuel_surcharge_cny,
                   service_abbrev, chargeable_weight_kg, rated_amount_cny
            FROM fedex_invoices 
            {where_clause}
            ORDER BY audit_timestamp DESC
        """
        cursor.execute(query, params)
        results = cursor.fetchall()
        conn.close()
        
        if not results:
            flash('No audit results found to export', 'warning')
            return redirect(url_for('fedex_invoices.batch_audit_results'))
        
        # Create detailed DataFrame
        detailed_data = []
        for row in results:
            # Row structure: invoice_no, awb_number, service_type, origin_country, dest_country, 
            #                actual_weight_kg, total_awb_amount_cny, expected_cost_cny, variance_cny, 
            #                audit_status, audit_timestamp, exchange_rate, fuel_surcharge_cny,
            #                service_abbrev, chargeable_weight_kg, rated_amount_cny
            
            # Calculate variance percentage
            variance_pct = (row[8] / row[6] * 100) if row[6] and row[6] != 0 else 0
            
            # Get zone from rate card lookup (simplified)
            zone = "Unknown"  # Would need zone lookup, but keeping simple for now
            rate_usd = "N/A"  # Would need rate card lookup
            
            detailed_data.append({
                'Invoice No': row[0],
                'AWB Number': row[1], 
                'Service Type': row[2],
                'Origin': row[3],
                'Destination': row[4],
                'Weight (kg)': row[5],
                'Chargeable Weight (kg)': row[14],  # chargeable_weight_kg
                'Invoiced Amount (CNY)': row[6],
                'Expected Amount (CNY)': row[7],
                'Variance (CNY)': row[8],
                'Variance %': f"{variance_pct:.2f}%",
                'Status': row[9],
                'Audit Date': row[10],
                'Exchange Rate': row[11] if row[11] else "N/A",  # exchange_rate
                'Fuel Surcharge (CNY)': row[12] if row[12] else 0,  # fuel_surcharge_cny
                'Service Code': row[13] if row[13] else row[2],  # service_abbrev or service_type
                'Base Rate (CNY)': row[15] if row[15] else 0,  # rated_amount_cny
                'Zone': zone,
                'Rate USD': rate_usd,
                'Audit Notes': f"System audit - {row[9]}"
            })
        
        # Create Excel file with detailed results
        import pandas as pd
        from io import BytesIO
        from flask import send_file
        
        detailed_df = pd.DataFrame(detailed_data)
        
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            detailed_df.to_excel(writer, sheet_name='Detailed Audit Results', index=False)
            
            # Auto-fit columns for better readability
            def auto_fit_columns(worksheet):
                from openpyxl.styles import Font, PatternFill, Alignment
                
                # Style the header row
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    # Style header cell
                    header_cell = worksheet[f"{column_letter}1"]
                    header_cell.font = header_font
                    header_cell.fill = header_fill
                    header_cell.alignment = header_alignment
                    
                    # Calculate column width
                    for cell in column:
                        try:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                        except:
                            pass
                    adjusted_width = min(max_length + 3, 60)  # Add padding and max width
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Apply auto-fit to the detailed results sheet
            auto_fit_columns(writer.sheets['Detailed Audit Results'])
        
        buffer.seek(0)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'fedex_detailed_audit_export_{timestamp}.xlsx'
        
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f'Error exporting detailed results: {str(e)}', 'error')
        return redirect(url_for('fedex_invoices.batch_audit_results'))
