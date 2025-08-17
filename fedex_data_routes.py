#!/usr/bin/env python3
"""
FedEx Data Management Routes
===========================

Flask routes for managing FedEx rate cards, zones, and surcharges.
"""

from flask import Blueprint, render_template, request, jsonify, flash, redirect, url_for, send_file, Response
import sqlite3
import json
from datetime import datetime
from fedex_rate_card_schema import FedExRateCardSchema

try:
    from auth_routes import require_auth, require_auth_api
except Exception:
    def require_auth(f): return f
    def require_auth_api(f): return f

fedex_data_bp = Blueprint('fedex_data', __name__)

@fedex_data_bp.route('/fedex/data-management')
@require_auth
def data_management(user_data=None):
    """Main data management dashboard"""
    
    with sqlite3.connect('fedex_audit.db') as conn:
        cursor = conn.cursor()
        
        # Get statistics
        stats = {}
        
        # Zone regions count
        cursor.execute("SELECT COUNT(*) FROM fedex_zone_regions WHERE active = 1")
        stats['zone_regions'] = cursor.fetchone()[0]
        
        # Country zones count
        cursor.execute("SELECT COUNT(*) FROM fedex_country_zones WHERE active = 1")
        stats['country_zones'] = cursor.fetchone()[0]
        
        # Rate cards count
        cursor.execute("SELECT COUNT(*) FROM fedex_rate_cards")
        stats['rate_cards'] = cursor.fetchone()[0]
        
        # Service types count
        cursor.execute("SELECT COUNT(*) FROM fedex_service_types WHERE active = 1")
        stats['service_types'] = cursor.fetchone()[0]
        
        # Surcharges count
        cursor.execute("SELECT COUNT(*) FROM fedex_surcharges WHERE active = 1")
        stats['surcharges'] = cursor.fetchone()[0]
        
        # Recent uploads
        cursor.execute('''
            SELECT filename, upload_date, status, records_processed
            FROM fedex_rate_card_uploads 
            ORDER BY upload_date DESC 
            LIMIT 5
        ''')
        recent_uploads = [
            {
                'filename': row[0],
                'upload_date': row[1],
                'status': row[2],
                'records_processed': row[3]
            }
            for row in cursor.fetchall()
        ]
        
        return render_template('fedex_data_management.html', 
                             stats=stats, 
                             recent_uploads=recent_uploads)

@fedex_data_bp.route('/fedex/zones')
@require_auth
def zones_management(user_data=None):
    """Zone management page"""
    
    with sqlite3.connect('fedex_audit.db') as conn:
        cursor = conn.cursor()
        
        # Get zone regions
        cursor.execute('''
            SELECT id, region_code, region_name, description, active, created_timestamp
            FROM fedex_zone_regions
            ORDER BY region_code
        ''')
        zone_regions = [
            {
                'id': row[0],
                'region_code': row[1],
                'region_name': row[2],
                'description': row[3],
                'active': bool(row[4]),
                'created_timestamp': row[5]
            }
            for row in cursor.fetchall()
        ]
        
        # Get zone matrix
        cursor.execute('''
            SELECT zm.id, zm.origin_country, zm.destination_region, zm.zone_letter, 
                   zm.active, zm.created_timestamp
            FROM fedex_zone_matrix zm
            ORDER BY zm.origin_country, zm.destination_region
        ''')
        zone_matrix = [
            {
                'id': row[0],
                'origin_country': row[1],
                'destination_region': row[2],
                'zone_letter': row[3],
                'active': bool(row[4]),
                'created_timestamp': row[5],
                'origin_name': row[1],  # Use origin_country as display name
                'dest_name': row[2]     # Use destination_region as display name
            }
            for row in cursor.fetchall()
        ]
        
        return render_template('fedex_zones.html', 
                             zone_regions=zone_regions, 
                             zone_matrix=zone_matrix)

@fedex_data_bp.route('/fedex/countries')
@require_auth
def countries_management(user_data=None):
    """Country zone mappings management"""
    
    with sqlite3.connect('fedex_audit.db') as conn:
        cursor = conn.cursor()
        
        # Get country zones with region names
        cursor.execute('''
            SELECT cz.id, cz.country_code, cz.country_name, cz.region_code, 
                   cz.sub_region, cz.zone_letter, cz.currency_code, cz.exchange_rate,
                   cz.active, cz.created_timestamp, zr.region_name
            FROM fedex_country_zones cz
            LEFT JOIN fedex_zone_regions zr ON cz.region_code = zr.region_code
            ORDER BY cz.country_name
        ''')
        countries = [
            {
                'id': row[0],
                'country_code': row[1],
                'country_name': row[2],
                'region_code': row[3],
                'sub_region': row[4],
                'zone_letter': row[5],
                'currency_code': row[6],
                'exchange_rate': row[7],
                'active': bool(row[8]),
                'created_timestamp': row[9],
                'region_name': row[10]
            }
            for row in cursor.fetchall()
        ]
        
        # Get available regions for dropdowns
        cursor.execute("SELECT region_code, region_name FROM fedex_zone_regions WHERE active = 1")
        regions = [{'code': row[0], 'name': row[1]} for row in cursor.fetchall()]
        
        return render_template('fedex_countries.html', 
                             countries=countries, 
                             regions=regions)

@fedex_data_bp.route('/fedex/rates')
@require_auth
def rates_management(user_data=None):
    """Rate cards management"""
    
    with sqlite3.connect('fedex_audit.db') as conn:
        cursor = conn.cursor()
        
        # Get total count
        cursor.execute("SELECT COUNT(*) FROM fedex_rate_cards")
        total_rates = cursor.fetchone()[0]
        
        # Get all rate cards (DataTables handles pagination client-side)
        cursor.execute('''
            SELECT id, rate_card_name, service_type, origin_region, destination_region,
                   zone_code, weight_from, weight_to, rate_usd, rate_type,
                   effective_date, expiry_date, created_timestamp
            FROM fedex_rate_cards 
            ORDER BY rate_card_name, weight_from
        ''')
        
        rate_cards = [
            {
                'id': row[0],
                'rate_card_name': row[1],
                'service_type': row[2],
                'origin_region': row[3],
                'destination_region': row[4],
                'zone_code': row[5],
                'weight_from': row[6],
                'weight_to': row[7],
                'rate_usd': row[8],
                'rate_type': row[9],
                'effective_date': row[10],
                'expiry_date': row[11],
                'created_timestamp': row[12]
            }
            for row in cursor.fetchall()
        ]
        
        # Get unique services for filter
        cursor.execute('SELECT DISTINCT service_type FROM fedex_rate_cards ORDER BY service_type')
        services = [{'service_code': row[0], 'service_name': row[0]} for row in cursor.fetchall()]
        
        # Get unique zones for filter  
        cursor.execute('SELECT DISTINCT zone_code FROM fedex_rate_cards ORDER BY zone_code')
        zones = [row[0] for row in cursor.fetchall()]
        
        return render_template('fedex_rates.html', 
                             rate_cards=rate_cards,
                             services=services,
                             zones=zones,
                             total_rates=total_rates)

@fedex_data_bp.route('/fedex/rates/export')
@require_auth
def export_rates(user_data=None):
    """Export rate cards to Excel or CSV.
    Query param: format=xlsx|csv (default xlsx)
    """
    fmt = request.args.get('format', 'xlsx').lower()
    # Fetch data
    with sqlite3.connect('fedex_audit.db') as conn:
        cursor = conn.cursor()
        cursor.execute('''
            SELECT rate_card_name, service_type, origin_region, destination_region,
                   zone_code, weight_from, IFNULL(weight_to, ''), rate_type, rate_usd, effective_date
            FROM fedex_rate_cards 
            ORDER BY rate_card_name, weight_from
        ''')
        rows = cursor.fetchall()

    headers = [
        'Rate Card', 'Service', 'Origin Region', 'Destination Region', 'Zone',
        'Weight From (kg)', 'Weight To (kg)', 'Rate Type', 'Rate (USD)', 'Effective Date'
    ]

    if fmt == 'csv':
        import csv
        from io import StringIO
        sio = StringIO()
        writer = csv.writer(sio)
        writer.writerow(headers)
        for r in rows:
            writer.writerow(r)
        data = sio.getvalue()
        sio.close()
        return Response(
            data,
            mimetype='text/csv',
            headers={
                'Content-Disposition': f"attachment; filename=fedex_rate_cards_{datetime.now().date()}.csv"
            }
        )

    # Try to create a real XLSX via openpyxl, else fall back to HTML .xls
    try:
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = 'FedEx Rate Cards'

        # Write header
        ws.append(headers)
        # Style header (basic bold)
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)

        # Write rows
        for r in rows:
            ws.append(list(r))

        # Auto-fit widths by max text length per column (cap at 50)
        for col_idx in range(1, len(headers) + 1):
            max_len = len(str(headers[col_idx - 1]))
            for row_idx in range(2, ws.max_row + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is None:
                    continue
                l = len(str(val))
                if l > max_len:
                    max_len = l
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return send_file(
            bio,
            as_attachment=True,
            download_name=f"fedex_rate_cards_{datetime.now().date()}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception:
        # Fallback: Excel-compatible HTML table with .xls extension
        from io import BytesIO
        def esc(x):
            s = '' if x is None else str(x)
            return s.replace('&', '&amp;').replace('<', '&lt;')
        thead = ''.join(f'<th>{esc(h)}</th>' for h in headers)
        tbody = ''.join('<tr>' + ''.join(f'<td>{esc(c)}</td>' for c in row) + '</tr>' for row in rows)
        html = f"""
        <html><head><meta charset='utf-8'>
        <style>table{{border-collapse:collapse}} td,th{{border:1px solid #ccc;padding:4px;white-space:nowrap;}}</style>
        </head><body><table><thead><tr>{thead}</tr></thead><tbody>{tbody}</tbody></table></body></html>
        """.strip()
        bio = BytesIO(html.encode('utf-8'))
        return send_file(
            bio,
            as_attachment=True,
            download_name=f"fedex_rate_cards_{datetime.now().date()}.xls",
            mimetype='application/vnd.ms-excel'
        )

@fedex_data_bp.route('/fedex/surcharges')
@require_auth
def surcharges_management(user_data=None):
    """Surcharges management"""
    
    with sqlite3.connect('fedex_audit.db') as conn:
        cursor = conn.cursor()
        
        # Get surcharges
        cursor.execute('''
            SELECT id, surcharge_code, surcharge_name, surcharge_description,
                   rate_type, rate_value, minimum_charge, maximum_charge,
                   applies_to_service, weight_threshold, active,
                   effective_date, expiry_date, created_timestamp
            FROM fedex_surcharges 
            ORDER BY surcharge_name
        ''')
        
        surcharges = [
            {
                'id': row[0],
                'surcharge_code': row[1],
                'surcharge_name': row[2],
                'surcharge_description': row[3],
                'rate_type': row[4],
                'rate_value': row[5],
                'minimum_charge': row[6],
                'maximum_charge': row[7],
                'applies_to_service': row[8],
                'weight_threshold': row[9],
                'active': bool(row[10]),
                'effective_date': row[11],
                'expiry_date': row[12],
                'created_timestamp': row[13]
            }
            for row in cursor.fetchall()
        ]
        
        return render_template('fedex_surcharges.html', surcharges=surcharges)

@fedex_data_bp.route('/fedex/api/initialize-schema', methods=['POST'])
@require_auth_api
def initialize_schema(user_data=None):
    """Initialize FedEx rate card schema"""
    try:
        schema = FedExRateCardSchema()
        schema.create_all_tables()
        
        return jsonify({
            'success': True,
            'message': 'FedEx rate card schema initialized successfully'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@fedex_data_bp.route('/fedex/api/stats')
@require_auth_api
def get_stats(user_data=None):
    """Get FedEx data statistics"""
    try:
        with sqlite3.connect('fedex_audit.db') as conn:
            cursor = conn.cursor()
            
            stats = {}
            
            # Get counts for each table
            tables = [
                ('zone_regions', 'fedex_zone_regions'),
                ('country_zones', 'fedex_country_zones'),
                ('rate_cards', 'fedex_rate_cards'),
                ('service_types', 'fedex_service_types'),
                ('surcharges', 'fedex_surcharges'),
                ('uploads', 'fedex_rate_card_uploads')
            ]
            
            for stat_name, table_name in tables:
                cursor.execute(f"SELECT COUNT(*) FROM {table_name}")
                stats[stat_name] = cursor.fetchone()[0]
            
            return jsonify({
                'success': True,
                'stats': stats
            })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500
