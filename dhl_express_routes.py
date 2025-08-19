#!/usr/bin/env python3
"""
DHL Express Invoice Audit Engine Routes
======================================

Flask routes for DHL Express invoice processing and audit functionality.
"""

from flask import Blueprint, request, jsonify, render_template, redirect, url_for, Response, make_response, flash, send_file
from werkzeug.utils import secure_filename
import os
import json
import sqlite3
import pandas as pd
from io import BytesIO
from datetime import datetime
from dhl_express_audit_engine import DHLExpressAuditEngine
# China audit engine (for CN invoices)
try:
    from dhl_express_china_audit_engine import DHLExpressChinaAuditEngine
except ImportError:
    DHLExpressChinaAuditEngine = None

# Import authentication
try:
    from auth_routes import require_auth, require_auth_api
except ImportError:
    # Fallback if auth is not available
    def require_auth(f):
        return f
    def require_auth_api(f):
        return f

dhl_express_routes = Blueprint('dhl_express', __name__)

@dhl_express_routes.route('/dhl-express')
@require_auth
def dhl_express_dashboard(user_data=None):
    """DHL Express main dashboard"""
    engine = DHLExpressAuditEngine()
    summary = engine.get_invoice_summary()
    
    return render_template('dhl_express_dashboard.html', summary=summary)

@dhl_express_routes.route('/dhl-express/upload', methods=['GET', 'POST'])
@require_auth
def upload_dhl_express_files(user_data=None):
    """Upload DHL Express invoices and rate cards with persistent status tracking"""
    if request.method == 'GET':
        return render_template('dhl_express_upload.html')
    
    try:
        from upload_status_manager import UploadStatusManager
        
        engine = DHLExpressAuditEngine()
        status_manager = UploadStatusManager(engine.db_path)
        
        # Prepare file information for session creation
        files_info = []
        uploaded_files = []
        
        for file_key in request.files:
            file = request.files[file_key]
            if file and file.filename:
                filename = secure_filename(file.filename)
                file_path = os.path.join('uploads', filename)
                
                # Ensure uploads directory exists
                os.makedirs('uploads', exist_ok=True)
                file.save(file_path)
                
                file_info = {
                    'filename': filename,
                    'original_filename': file.filename,
                    'size': os.path.getsize(file_path)
                }
                
                files_info.append(file_info)
                uploaded_files.append((file_path, filename, file.filename))
        
        if not files_info:
            return jsonify({
                'success': False,
                'error': 'No files uploaded'
            }), 400
        
        # Create upload session
        session_id = status_manager.create_upload_session(files_info)
        
        # Process uploaded files
        results = {'success': True, 'files_processed': [], 'session_id': session_id}
        
        for file_path, filename, original_filename in uploaded_files:
            try:
                # Process based on file type
                if filename.lower().endswith('.csv'):
                    # Process as invoice CSV
                    result = engine.load_invoices_from_csv(file_path)
                    
                    file_result = {
                        'filename': filename,
                        'type': 'invoice_csv',
                        'result': result
                    }
                    
                elif filename.lower().endswith(('.xlsx', '.xls')):
                    # Detect file type by checking sheet names
                    try:
                        import openpyxl
                        wb = openpyxl.load_workbook(file_path, read_only=True)
                        sheet_names = wb.sheetnames
                        wb.close()
                        
                        # Check if it's CN (Chinese) International rate card
                        cn_international_sheets = ['CN TD Exp WW', 'CN TD Imp WW']
                        if any(sheet in sheet_names for sheet in cn_international_sheets):
                            # Process as Chinese International rate card
                            try:
                                from dhl_express_china_rate_card_loader import DHLExpressChinaRateCardLoader
                                
                                loader = DHLExpressChinaRateCardLoader()
                                result = loader.load_complete_rate_cards(file_path)
                                
                                if result['success']:
                                    file_result = {
                                        'filename': filename,
                                        'type': 'cn_international_rate_card',
                                        'result': {
                                            'success': True,
                                            'message': 'Chinese International rate card loaded successfully',
                                            'details': {
                                                'import_rates_loaded': result['import_rates_loaded'],
                                                'export_rates_loaded': result['export_rates_loaded'],
                                                'sections_processed': result['sections_processed'],
                                                'currency': 'CNY'
                                            },
                                            'records_processed': f"Import: {result['import_rates_loaded']}, Export: {result['export_rates_loaded']}"
                                        }
                                    }
                                else:
                                    file_result = {
                                        'filename': filename,
                                        'type': 'cn_international_rate_card',
                                        'result': {
                                            'success': False,
                                            'message': 'Chinese International rate card loading failed',
                                            'error': result.get('error', 'Unknown error')
                                        }
                                    }
                                    
                            except Exception as cn_rate_error:
                                file_result = {
                                    'filename': filename,
                                    'type': 'cn_international_rate_card',
                                    'result': {
                                        'success': False,
                                        'message': f'Chinese International rate card loading failed: {str(cn_rate_error)}',
                                        'error': str(cn_rate_error),
                                        'error_type': type(cn_rate_error).__name__
                                    }
                                }
                        
                        # Check if it's Chinese invoice file (has Table1 sheet with DHL invoice data)
                        elif 'Table1' in sheet_names:
                            # Check if it looks like a Chinese DHL invoice
                            try:
                                # Quick peek at the structure
                                test_df = pd.read_excel(file_path, sheet_name='Table1', nrows=1)
                                expected_columns = ['Air waybill', 'Invoice Number', 'BCU Total', 'LCU Total']
                                if all(col in test_df.columns for col in expected_columns):
                                    # Process as Chinese invoice
                                    from dhl_express_china_invoice_loader import DHLExpressChinaInvoiceLoader
                                    
                                    loader = DHLExpressChinaInvoiceLoader()
                                    result = loader.load_invoices_from_excel(file_path, 
                                                                            user_data.get('username', 'unknown') if user_data else 'system')
                                    
                                    if result['success']:
                                        file_result = {
                                            'filename': filename,
                                            'type': 'cn_invoice',
                                            'result': {
                                                'success': True,
                                                'message': 'Chinese DHL invoices loaded successfully',
                                                'details': {
                                                    'records_loaded': result['records_loaded'],
                                                    'total_records': result['total_records'],
                                                    'upload_id': result['upload_id']
                                                },
                                                'records_processed': result['records_loaded']
                                            }
                                        }
                                    else:
                                        file_result = {
                                            'filename': filename,
                                            'type': 'cn_invoice',
                                            'result': {
                                                'success': False,
                                                'message': 'Chinese DHL invoice loading failed',
                                                'error': result.get('error', 'Unknown error')
                                            }
                                        }
                                else:
                                    file_result = {
                                        'filename': filename,
                                        'type': 'table1_unknown',
                                        'result': {
                                            'success': False,
                                            'message': 'Table1 sheet found but does not appear to be a DHL invoice format',
                                            'error': f'Expected columns: {expected_columns}, Found: {list(test_df.columns)[:10]}'
                                        }
                                    }
                                    
                            except Exception as invoice_error:
                                file_result = {
                                    'filename': filename,
                                    'type': 'cn_invoice_error',
                                    'result': {
                                        'success': False,
                                        'message': f'Chinese invoice processing failed: {str(invoice_error)}',
                                        'error': str(invoice_error),
                                        'error_type': type(invoice_error).__name__
                                    }
                                }
                        
                        # Check if it's AU Domestic rate card (legacy support)
                        elif 'AU Zones TD Dom' in sheet_names or 'AU Matrix TD Dom' in sheet_names:
                            file_result = {
                                'filename': filename,
                                'type': 'au_domestic_rate_card',
                                'result': {
                                    'success': False,
                                    'message': 'AU Domestic rate cards are no longer supported. Please upload Chinese rate cards.',
                                    'error': 'This system now only supports Chinese DHL Express rate cards'
                                }
                            }
                        
                        # Check if it's AU International rate card (legacy support)
                        elif 'AU TD Exp WW' in sheet_names or 'AU TD Imp WW' in sheet_names:
                            file_result = {
                                'filename': filename,
                                'type': 'au_international_rate_card',
                                'result': {
                                    'success': False,
                                    'message': 'AU International rate cards are no longer supported. Please upload Chinese rate cards.',
                                    'error': 'This system now only supports Chinese DHL Express rate cards'
                                }
                            }
                        
                        else:
                            # Unknown Excel file type
                            file_result = {
                                'filename': filename,
                                'type': 'unknown_excel',
                                'result': {
                                    'success': False,
                                    'message': 'Unknown Excel file format. Expected Chinese DHL Express rate cards (CN TD Exp WW/CN TD Imp WW sheets) or Chinese invoices (Table1 sheet).',
                                    'error': f'Sheet names found: {", ".join(sheet_names)}',
                                    'expected_sheets': cn_international_sheets + ['Table1 (for invoices)']
                                }
                            }
                        
                    except Exception as detection_error:
                        file_result = {
                            'filename': filename,
                            'type': 'excel_error',
                            'result': {
                                'success': False,
                                'message': f'Excel file detection failed: {str(detection_error)}',
                                'error': str(detection_error),
                                'error_type': type(detection_error).__name__
                            }
                        }
                
                else:
                    file_result = {
                        'filename': filename,
                        'type': 'unknown',
                        'result': {
                            'success': False,
                            'message': 'Unsupported file type',
                            'error': 'Only CSV and Excel files are supported'
                        }
                    }
                
                # Update individual file status
                status_manager.update_file_result(session_id, filename, file_result['result'])
                results['files_processed'].append(file_result)
                
            except Exception as file_error:
                error_result = {
                    'success': False,
                    'message': f'Error processing file: {str(file_error)}',
                    'error': str(file_error),
                    'error_type': type(file_error).__name__
                }
                
                file_result = {
                    'filename': filename,
                    'type': 'error',
                    'result': error_result
                }
                
                status_manager.update_file_result(session_id, filename, error_result)
                results['files_processed'].append(file_result)
        
        # Check if any files failed
        failed_files = [f for f in results['files_processed'] if not f['result'].get('success', True)]
        if failed_files:
            results['success'] = False
            results['error'] = f"Failed to process {len(failed_files)} file(s)"
            results['error_details'] = failed_files
        
        # Finalize upload session
        final_status = status_manager.finalize_upload_session(session_id, results)
        results['final_status'] = final_status
        
        return jsonify(results)
        
    except Exception as e:
        import traceback
        error_result = {
            'success': False, 
            'error': str(e),
            'error_details': {
                'exception_type': type(e).__name__,
                'traceback': traceback.format_exc()
            }
        }
        
        # Try to update session if it was created
        try:
            if 'session_id' in locals():
                status_manager.finalize_upload_session(session_id, error_result)
        except:
            pass  # Ignore secondary errors
        
        return jsonify(error_result), 500

@dhl_express_routes.route('/dhl-express/dashboard')
@require_auth
def upload_dashboard(user_data=None):
    """Upload monitoring page for status and history tracking"""
    return render_template('upload_dashboard.html')

@dhl_express_routes.route('/dhl-express/api/upload-status')
@require_auth
def get_upload_status(user_data=None):
    """Get current persistent upload status"""
    try:
        from upload_status_manager import UploadStatusManager
        
        engine = DHLExpressAuditEngine()
        status_manager = UploadStatusManager(engine.db_path)
        
        current_status = status_manager.get_current_status()
        
        return jsonify({
            'success': True,
            'has_status': current_status is not None,
            'status': current_status
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@dhl_express_routes.route('/dhl-express/api/upload-history')
@require_auth
def get_upload_history(user_data=None):
    """Get upload history"""
    try:
        from upload_status_manager import UploadStatusManager
        
        engine = DHLExpressAuditEngine()
        status_manager = UploadStatusManager(engine.db_path)
        
        limit = request.args.get('limit', 20, type=int)
        history = status_manager.get_upload_history(limit)
        
        return jsonify({
            'success': True,
            'history': history
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@dhl_express_routes.route('/dhl-express/api/clear-status', methods=['POST'])
@require_auth
def clear_upload_status(user_data=None):
    """Clear current upload status"""
    try:
        from upload_status_manager import UploadStatusManager
        
        engine = DHLExpressAuditEngine()
        status_manager = UploadStatusManager(engine.db_path)
        
        status_manager.clear_current_status()
        
        return jsonify({
            'success': True,
            'message': 'Upload status cleared successfully'
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@dhl_express_routes.route('/dhl-express/api/session-details/<session_id>')
@require_auth
def get_session_details(session_id, user_data=None):
    """Get detailed information about a specific upload session"""
    try:
        from upload_status_manager import UploadStatusManager
        
        engine = DHLExpressAuditEngine()
        status_manager = UploadStatusManager(engine.db_path)
        
        details = status_manager.get_session_details(session_id)
        
        if details:
            return jsonify({
                'success': True,
                'details': details
            })
        else:
            return jsonify({
                'success': False,
                'error': 'Session not found'
            }), 404
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@dhl_express_routes.route('/dhl-express/invoices')
@require_auth
def list_dhl_express_invoices(user_data=None):
    """List all DHL Express invoices"""
    engine = DHLExpressAuditEngine()
    
    # Get invoices from database (now using Chinese invoice table)
    import sqlite3
    conn = sqlite3.connect(engine.db_path)
    cursor = conn.cursor()
    
    # Check if we have Chinese invoices
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='dhl_express_china_invoices'")
    china_table_exists = cursor.fetchone() is not None
    
    if china_table_exists:
        # Use Chinese invoice table
        cursor.execute('''
            SELECT DISTINCT invoice_number, invoice_date, bill_to_account_name, 
                   COUNT(*) as line_count, SUM(lcu_total) as total_amount,
                   MIN(created_timestamp) as loaded_date,
                   local_currency, billing_currency
            FROM dhl_express_china_invoices
            GROUP BY invoice_number, invoice_date, bill_to_account_name
            ORDER BY invoice_date DESC
        ''')
        
        invoices = cursor.fetchall()
        
        # Convert to list of dictionaries for Chinese invoices
        invoice_list = []
        for inv in invoices:
            invoice_list.append({
                'invoice_no': inv[0],
                'invoice_date': inv[1] if inv[1] else 'N/A',
                'company_name': inv[2] if inv[2] else 'N/A',
                'line_count': inv[3],
                'total_amount': round(inv[4], 2) if inv[4] else 0,
                'loaded_date': inv[5] if inv[5] else 'N/A',
                'currency': inv[6] if inv[6] else inv[7],  # Local currency or billing currency
                'is_chinese': True
            })
            
    else:
        # Fallback to old AU invoice table (if any legacy data exists)
        cursor.execute('''
            SELECT DISTINCT invoice_no, invoice_date, company_name, 
                   COUNT(*) as line_count, SUM(amount) as total_amount,
                   MIN(created_timestamp) as loaded_date
            FROM dhl_express_invoices
            GROUP BY invoice_no, invoice_date, company_name
            ORDER BY invoice_date DESC
        ''')
        
        invoices = cursor.fetchall()
        
        # Convert to list of dictionaries for AU invoices
        invoice_list = []
        for inv in invoices:
            invoice_list.append({
                'invoice_no': inv[0],
                'invoice_date': inv[1],
                'company_name': inv[2],
                'line_count': inv[3],
                'total_amount': round(inv[4], 2),
                'loaded_date': inv[5],
                'currency': 'AUD',
                'is_chinese': False
            })
    
    conn.close()
    
    return render_template('dhl_express_invoices.html', invoices=invoice_list, is_chinese_system=china_table_exists)

@dhl_express_routes.route('/dhl-express/invoice/<invoice_no>')
@require_auth
def view_dhl_express_invoice(invoice_no, user_data=None):
    """View detailed DHL Express invoice"""
    engine = DHLExpressAuditEngine()
    
    # Get invoice details
    import sqlite3
    conn = sqlite3.connect(engine.db_path)
    cursor = conn.cursor()
    
    # Check if we have Chinese invoices
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='dhl_express_china_invoices'")
    china_table_exists = cursor.fetchone() is not None
    
    if china_table_exists:
        # Try Chinese invoice table first
        cursor.execute('''
            SELECT * FROM dhl_express_china_invoices
            WHERE invoice_number = ?
            ORDER BY id
        ''', (invoice_no,))
        
        rows = cursor.fetchall()
        
        if rows:
            # Get column names for Chinese invoice table
            cursor.execute("PRAGMA table_info(dhl_express_china_invoices)")
            columns = [col[1] for col in cursor.fetchall()]
            
            # Convert to dictionaries
            invoice_lines = []
            for row in rows:
                line_dict = dict(zip(columns, row))
                invoice_lines.append(line_dict)
            
            # Get invoice header info from first line
            first_line = invoice_lines[0]
            invoice_header = {
                'invoice_no': first_line['invoice_number'],
                'invoice_date': first_line['invoice_date'] or 'N/A',
                'company_name': first_line['bill_to_account_name'] or 'N/A',
                'account_number': first_line['bill_to_account'],
                'currency': first_line['local_currency'] or first_line['billing_currency'],
                'total_bcu': sum(line.get('bcu_total', 0) or 0 for line in invoice_lines),
                'total_lcu': sum(line.get('lcu_total', 0) or 0 for line in invoice_lines),
                'is_chinese': True
            }
            
            conn.close()
            return render_template('dhl_express_invoice_detail.html', 
                                 invoice_header=invoice_header, 
                                 invoice_lines=invoice_lines,
                                 is_chinese=True)
    
    # Fallback to AU invoice table
    def parse_address_details(address_string):
        """Parse semicolon-separated address details"""
        if not address_string:
            return {}
        
        parts = [part.strip() for part in address_string.split(';') if part.strip()]
        
        # Standard DHL format: Company;Address1;Address2;City;City2;State;PostalCode;Country;Contact;Email
        parsed = {
            'company': parts[0] if len(parts) > 0 else '',
            'address1': parts[1] if len(parts) > 1 else '',
            'address2': parts[2] if len(parts) > 2 else '',
            'city': parts[3] if len(parts) > 3 else '',
            'city2': parts[4] if len(parts) > 4 else '',
            'state': parts[5] if len(parts) > 5 else '',
            'postal_code': parts[6] if len(parts) > 6 else '',
            'country': parts[7] if len(parts) > 7 else '',
            'contact': parts[8] if len(parts) > 8 else '',
            'email': parts[9] if len(parts) > 9 else ''
        }
        return parsed
    
    # Get AU invoice header
    cursor.execute('''
        SELECT DISTINCT invoice_no, invoice_date, company_name, account_number,
               shipper_details, receiver_details
        FROM dhl_express_invoices
        WHERE invoice_no = ?
    ''', (invoice_no,))
    
    header = cursor.fetchone()
    if not header:
        conn.close()
        return "Invoice not found", 404
    
    # Get AU invoice lines  
    cursor.execute('''
        SELECT line_number, dhl_product_description, amount, weight_charge,
               discount_amount, tax_amount, awb_number, weight, origin_code,
               destination_code, shipment_date
        FROM dhl_express_invoices
        WHERE invoice_no = ?
        ORDER BY line_number
    ''', (invoice_no,))
    
    lines = cursor.fetchall()
    conn.close()
    
    # Convert to dictionaries for AU invoices
    invoice_header = {
        'invoice_no': header[0],
        'invoice_date': header[1],
        'company_name': header[2],
        'account_number': header[3],
        'shipper': parse_address_details(header[4]),
        'receiver': parse_address_details(header[5]),
        'currency': 'AUD',
        'is_chinese': False
    }
    
    invoice_lines = []
    for line in lines:
        # Parse and reformat shipment_date from database if it exists
        shipment_date_display = line[10]  # Get original date string
        if shipment_date_display:
            try:
                # Try to parse the date and convert it to proper US format display
                from datetime import datetime
                # If the date is in YYYY-MM-DD format but was originally MM-DD-YYYY
                if '-' in shipment_date_display and len(shipment_date_display) == 10:
                    # Parse as YYYY-MM-DD and reinterpret as MM-DD-YYYY
                    date_parts = shipment_date_display.split('-')
                    if len(date_parts) == 3:
                        year, month, day = date_parts
                        # Reinterpret: if it looks like 2025-01-04, it should be Jan 4, 2025
                        # But if it looks like 2025-09-04, it should be Jan 8, 2025 (month 09 = day 08)
                        # Based on your indication: 01-04 = Jan 4, 09-04 = Jan 8 (so 09 is actually 08 day)
                        shipment_date_display = f"{month}-{day}-{year}"
            except:
                # If parsing fails, keep original
                pass
        
        invoice_lines.append({
            'line_number': line[0],
            'product_description': line[1],
            'amount': line[2],
            'weight_charge': line[3],
            'discount_amount': line[4],
            'tax_amount': line[5],
            'awb_number': line[6],
            'weight': line[7],
            'origin_code': line[8],
            'destination_code': line[9],
            'shipment_date': shipment_date_display
        })
    
    # Check if invoice image exists for this invoice number
    import sqlite3
    conn_img = sqlite3.connect('dhl_audit.db')
    cursor_img = conn_img.cursor()
    cursor_img.execute('SELECT COUNT(*) FROM invoice_images WHERE invoice_number = ? AND invoice_type = ?', 
                   (invoice_no, 'DHL_EXPRESS'))
    has_invoice_image = cursor_img.fetchone()[0] > 0
    conn_img.close()
    
    return render_template('dhl_express_invoice_detail.html', 
                         invoice_header=invoice_header, 
                         invoice_lines=invoice_lines, 
                         has_invoice_image=has_invoice_image,
                         is_chinese=False)

@dhl_express_routes.route('/dhl-express/audit/<invoice_no>')
@require_auth
def audit_dhl_express_invoice(invoice_no, user_data=None):
    """Audit a specific DHL Express invoice"""
    # Let's create a minimal HTML page with JavaScript to handle the data
    engine = DHLExpressAuditEngine()
    
    try:
        audit_result = engine.audit_invoice(invoice_no)
        
        # Create a simple HTML page that will redirect to the proper template
        # Print audit result for debugging
        print(f"Audit result: {type(audit_result)}")
        print(f"Keys: {audit_result.keys() if isinstance(audit_result, dict) else 'Not a dict'}")
        
        # Parse line_items if it's a string
        if isinstance(audit_result.get('line_items', None), str):
            try:
                audit_result['line_items'] = json.loads(audit_result['line_items'])
                print(f"Successfully parsed line_items JSON")
            except Exception as e:
                print(f"Error parsing line_items JSON: {e}")
                # Try a less strict parsing
                import ast
                try:
                    audit_result['line_items'] = ast.literal_eval(audit_result['line_items'])
                    print(f"Parsed line_items with ast.literal_eval")
                except Exception as e2:
                    print(f"Failed to parse with ast too: {e2}")
                    audit_result['line_items'] = []
        
        # Ensure all required fields are present
        if 'line_items' not in audit_result or not audit_result['line_items']:
            print("No line_items found in audit result")
            audit_result['line_items'] = []
            
        # Set consistent field names
        audit_result['audit_status'] = audit_result.get('status', 'REVIEW')
        audit_result['detailed_results'] = audit_result.get('line_items', [])
        
        # Count failed lines and passed lines
        failed_lines = 0
        passed_lines = 0
        for item in audit_result.get('line_items', []):
            if item.get('result') == 'FAIL':
                failed_lines += 1
            elif item.get('result') == 'PASS':
                passed_lines += 1
                
        audit_result['lines_failed'] = failed_lines
        audit_result['lines_passed'] = passed_lines
        audit_result['lines_audited'] = len(audit_result.get('line_items', []))
        audit_result['confidence_score'] = float(audit_result.get('confidence', 70)) / 100 if float(audit_result.get('confidence', 70)) > 1 else float(audit_result.get('confidence', 0.7))
        
        # Calculate variance percentage if not present
        if 'variance_percentage' not in audit_result or not isinstance(audit_result.get('variance_percentage'), (int, float)):
            invoice_amount = float(audit_result.get('total_invoice_amount', 0))
            expected_amount = float(audit_result.get('total_expected_amount', 0))
            variance = float(audit_result.get('total_variance', 0))
            
            # Calculate variance percentage safely
            if invoice_amount > 0:
                audit_result['variance_percentage'] = (variance / invoice_amount) * 100
            else:
                audit_result['variance_percentage'] = 0.0
        
        # Add safety check to ensure we have required fields
        required_fields = ['invoice_no', 'total_invoice_amount', 'total_expected_amount', 'total_variance']
        for field in required_fields:
            if field not in audit_result:
                audit_result[field] = 0 if field != 'invoice_no' else invoice_no
        
        # Check if invoice image exists for this invoice number
        import sqlite3
        conn = sqlite3.connect('dhl_audit.db')
        cursor = conn.cursor()
        cursor.execute('SELECT COUNT(*) FROM invoice_images WHERE invoice_number = ? AND invoice_type = ?', 
                       (invoice_no, 'DHL_EXPRESS'))
        has_invoice_image = cursor.fetchone()[0] > 0
        conn.close()
                
        # Return content with explicit Content-Type header
        response = make_response(render_template(
            'dhl_express_audit_results.html',
            audit=audit_result,
            has_invoice_image=has_invoice_image
        ))
        response.headers['Content-Type'] = 'text/html; charset=utf-8'
        return response
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Error in audit_dhl_express_invoice: {error_details}")
        
        # More specific user-friendly error message
        error_message = str(e)
        user_message = "There was an error processing the audit."
        
        if "unsupported format string" in error_message:
            user_message = "Error in formatting audit data. This might be due to missing or invalid values."
        elif "JSON" in error_message or "json" in error_message:
            user_message = "Error parsing audit data. The format may be invalid."
        elif "KeyError" in error_message:
            user_message = "Required data field is missing from the audit results."
        
        # Return a user-friendly error page
        return render_template('error.html',
                             error=error_message,
                             title="Audit Error",
                             message=user_message), 500



@dhl_express_routes.route('/api/dhl-express/audit/<invoice_no>')
@require_auth_api
def api_audit_dhl_express_invoice(invoice_no, user_data=None):
    """API endpoint for DHL Express invoice audit"""
    engine = DHLExpressAuditEngine()
    
    try:
        result = engine.audit_invoice(invoice_no)
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@dhl_express_routes.route('/dhl-express/rate-cards')
@require_auth
def view_dhl_express_rate_cards(user_data=None):
    """View loaded DHL Express rate cards and all related tables"""
    engine = DHLExpressAuditEngine()
    
    import sqlite3
    conn = sqlite3.connect(engine.db_path)
    cursor = conn.cursor()
    
    # Get rate card summary
    cursor.execute('''
        SELECT service_type, rate_section,
               COUNT(*) as rate_count, MIN(weight_from) as min_weight,
               MAX(weight_to) as max_weight, MAX(created_timestamp) as effective_date
        FROM dhl_express_rate_cards
        GROUP BY service_type, rate_section
        ORDER BY service_type, rate_section
    ''')
    
    rate_cards = cursor.fetchall()
    
    # Get services summary with enhanced details
    cursor.execute('''
        SELECT COUNT(*) as total_services,
               SUM(CASE WHEN is_special_agreement = 1 THEN 1 ELSE 0 END) as special_services,
               SUM(CASE WHEN is_special_agreement = 0 THEN 1 ELSE 0 END) as published_services
        FROM dhl_express_services_surcharges
    ''')
    
    services_summary = cursor.fetchone()
    
    # Get detailed service charges
    cursor.execute('''
        SELECT service_code, service_name, charge_type, charge_amount,
               is_special_agreement, created_timestamp,
               0 as minimum_charge,
               0 as percentage_rate,
               COALESCE(products_applicable, 'All Products') as products_applicable
        FROM dhl_express_services_surcharges
        ORDER BY service_code
        LIMIT 50
    ''')
    
    service_charges = cursor.fetchall()
    
    # Get 3rd party rates summary
    cursor.execute('''
        SELECT 'International' as service_type, COUNT(*) as rate_count,
               MIN(weight_kg) as min_weight, MAX(weight_kg) as max_weight,
               MAX(created_at) as effective_date
        FROM dhl_express_3rd_party_rates
        WHERE weight_kg IS NOT NULL
        UNION ALL
        SELECT 'Domestic Matrix' as service_type, COUNT(*) as rate_count,
               0 as min_weight, 0 as max_weight,
               MAX(created_timestamp) as effective_date
        FROM dhl_express_3rd_party_domestic_matrix
        ORDER BY service_type
    ''')
    
    third_party_rates = cursor.fetchall()
    
    # Get zone mapping summary
    cursor.execute('''
        SELECT zone_number as zone, COUNT(*) as country_count
        FROM dhl_express_zone_mapping
        WHERE zone_number IS NOT NULL
        GROUP BY zone_number
        ORDER BY zone_number
    ''')
    
    zone_mappings = cursor.fetchall()
    
    # Get country codes summary
    cursor.execute('''
        SELECT COUNT(DISTINCT country_code) as total_countries,
               COUNT(DISTINCT country_name) as total_country_names
        FROM dhl_express_country_codes
        WHERE country_code IS NOT NULL
    ''')
    
    country_summary = cursor.fetchone()
    
    # Get sample countries
    cursor.execute('''
        SELECT DISTINCT country_name
        FROM dhl_express_country_codes
        WHERE country_name IS NOT NULL
        ORDER BY country_name
        LIMIT 20
    ''')
    
    sample_countries = cursor.fetchall()
    
    # Get multiplier ranges
    cursor.execute('''
        SELECT service_type, weight_from, weight_to, zone_5,
               created_timestamp
        FROM dhl_express_rate_cards
        WHERE is_multiplier = 1
        ORDER BY service_type, weight_from
    ''')

    multiplier_ranges = cursor.fetchall()
    
    # Get AU Domestic rate card data
    au_domestic_summary = {}
    try:
        # Check if AU domestic tables exist
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='dhl_express_au_domestic_zones'")
        if cursor.fetchone():
            # Get zones summary
            cursor.execute('SELECT COUNT(*) FROM dhl_express_au_domestic_zones')
            zones_count = cursor.fetchone()[0]
            
            # Get matrix summary  
            cursor.execute('SELECT COUNT(*) FROM dhl_express_au_domestic_matrix')
            matrix_count = cursor.fetchone()[0]
            
            # Get rates summary
            cursor.execute('SELECT COUNT(*) FROM dhl_express_au_domestic_rates')
            rates_count = cursor.fetchone()[0]
            
            # Get latest upload info
            cursor.execute('''
                SELECT filename, upload_date, uploaded_by, status
                FROM dhl_express_au_domestic_uploads 
                ORDER BY upload_date DESC LIMIT 1
            ''')
            latest_upload = cursor.fetchone()
            
            # Get sample zones
            cursor.execute('''
                SELECT zone_number, city_name, city_code
                FROM dhl_express_au_domestic_zones 
                WHERE city_name IS NOT NULL 
                ORDER BY zone_number LIMIT 5
            ''')
            sample_zones = cursor.fetchall()
            
            au_domestic_summary = {
                'zones_count': zones_count,
                'matrix_count': matrix_count,
                'rates_count': rates_count,
                'latest_upload': latest_upload,
                'sample_zones': sample_zones,
                'available': True
            }
        else:
            au_domestic_summary = {'available': False}
    except Exception as e:
        au_domestic_summary = {'available': False, 'error': str(e)}
    
    conn.close()    # Convert to dictionaries
    rate_card_list = []
    for rc in rate_cards:
        rate_card_list.append({
            'service_type': rc[0],
            'rate_section': rc[1],
            'rate_count': rc[2],
            'weight_range': f"{rc[3]}-{rc[4]} kg",
            'effective_date': rc[5]
        })
    
    services_info = {
        'total_services': services_summary[0] if services_summary else 0,
        'special_services': services_summary[1] if services_summary else 0,
        'published_services': services_summary[2] if services_summary else 0
    }
    
    # Convert service charges to list
    service_charges_list = []
    for sc in service_charges:
        service_charges_list.append({
            'service_code': sc[0],
            'service_name': sc[1],
            'charge_type': sc[2],
            'charge_amount': sc[3],
            'is_special': sc[4],
            'created_timestamp': sc[5],
            'minimum_charge': sc[6],
            'percentage_rate': sc[7],
            'products_applicable': sc[8]
        })
    
    # Convert 3rd party rates
    third_party_list = []
    for tp in third_party_rates:
        third_party_list.append({
            'service_type': tp[0],
            'rate_count': tp[1],
            'weight_range': f"{tp[2]}-{tp[3]} kg" if tp[2] > 0 and tp[3] > 0 else "Matrix Data",
            'effective_date': tp[4]
        })
    
    # Convert zone mappings
    zone_mappings_list = []
    for zm in zone_mappings:
        zone_mappings_list.append({
            'zone': zm[0],
            'country_count': zm[1]
        })
    
    country_info = {
        'total_countries': country_summary[0] if country_summary else 0,
        'total_regions': country_summary[1] if country_summary else 0
    }
    
    # Convert sample countries
    country_regions_list = []
    if sample_countries:
        # Group sample countries into chunks for display
        countries_text = ', '.join([c[0] for c in sample_countries])
        country_regions_list.append({
            'region': 'Sample Countries',
            'sample_countries': countries_text
        })
    
    # Convert multiplier ranges
    multiplier_ranges_list = []
    for mr in multiplier_ranges:
        multiplier_ranges_list.append({
            'service_type': mr[0],
            'weight_range': f"{mr[1]}-{mr[2]} kg",
            'zone_5_rate': mr[3],
            'created_timestamp': mr[4]
        })

    return render_template('dhl_express_rate_cards.html', 
                         rate_cards=rate_card_list, 
                         services=services_info,
                         service_charges=service_charges_list,
                         third_party_rates=third_party_list,
                         zone_mappings=zone_mappings_list,
                         country_info=country_info,
                         country_regions=country_regions_list,
                         multiplier_ranges=multiplier_ranges_list,
                         au_domestic=au_domestic_summary)


@dhl_express_routes.route('/dhl-express/rate-card-details', methods=['GET'])
@require_auth
def get_rate_card_details(user_data=None):
    """Get detailed rate card information for a specific service type and rate section"""
    from flask import jsonify, request
    
    # Initialize the engine to get the correct database path
    engine = DHLExpressAuditEngine()
    
    service_type = request.args.get('service_type')
    rate_section = request.args.get('rate_section')
    
    if not service_type or not rate_section:
        return jsonify({
            'error': 'Missing required parameters: service_type and rate_section'
        }), 400
    
    try:
        conn = sqlite3.connect(engine.db_path)
        cursor = conn.cursor()
        
        # Check if the rate_cards table exists and print schema
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='dhl_express_rate_cards'")
        table_exists = cursor.fetchone()
        
        if not table_exists:
            conn.close()
            return jsonify({
                'error': 'Rate cards table does not exist in database'
            }), 500
            
        # Get table schema to verify column names
        cursor.execute("PRAGMA table_info(dhl_express_rate_cards)")
        columns = cursor.fetchall()
        column_names = [col[1] for col in columns]
        
        # Fetch the rate card entries for the specified service type and rate section
        try:
            cursor.execute('''
                SELECT weight_from, weight_to, zone_1, zone_2, zone_3, zone_4,
                      zone_5, zone_6, zone_7, zone_8, zone_9, is_multiplier
                FROM dhl_express_rate_cards
                WHERE service_type = ? AND rate_section = ?
                ORDER BY weight_from
            ''', (service_type, rate_section))
            
            rows = cursor.fetchall()
            
            if not rows:
                conn.close()
                return jsonify({
                    'error': 'No rate card entries found for the specified parameters',
                    'available_columns': column_names
                }), 404
        except sqlite3.OperationalError as e:
            conn.close()
            return jsonify({
                'error': f'Database error: {str(e)}',
                'available_columns': column_names
            }), 500
        
        # Convert rows to list of dictionaries
        rate_entries = []
        for row in rows:
            entry = {
                'weight_from': row[0],
                'weight_to': row[1],
                'zone_1': row[2],
                'zone_2': row[3],
                'zone_3': row[4],
                'zone_4': row[5],
                'zone_5': row[6],
                'zone_6': row[7],
                'zone_7': row[8],
                'zone_8': row[9],
                'zone_9': row[10],
                'is_multiplier': bool(row[11])
            }
            rate_entries.append(entry)
        
        # Get zone descriptions
        zones_info = {}
        
        # Check if zone mapping table exists
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='dhl_express_zone_mapping'")
        zone_table_exists = cursor.fetchone()
        
        if zone_table_exists:
            try:
                cursor.execute('''
                    SELECT zone_number, origin_code, destination_code
                    FROM dhl_express_zone_mapping
                    WHERE service_type = ?
                    ORDER BY zone_number
                ''', (service_type,))
                
                zone_mappings = cursor.fetchall()
                
                for mapping in zone_mappings:
                    zone = f"zone_{mapping[0]}"
                    if zone not in zones_info:
                        zones_info[zone] = []
                    zones_info[zone].append({
                        'origin': mapping[1],
                        'destination': mapping[2]
                    })
            except sqlite3.OperationalError as e:
                # Log the error but continue - zone mappings are optional
                zones_info = {'error': f'Failed to load zone mappings: {str(e)}'}
        
        conn.close()
        return jsonify({
            'service_type': service_type,
            'rate_section': rate_section,
            'rate_entries': rate_entries,
            'zones_info': zones_info
        })
    except Exception as e:
        return jsonify({
            'error': f'Error retrieving rate card details: {str(e)}'
        }), 500

@dhl_express_routes.route('/dhl-express/audit-summary')
@require_auth
def dhl_express_audit_summary(user_data=None):
    """DHL Express audit summary dashboard"""
    engine = DHLExpressAuditEngine()
    
    import sqlite3
    conn = sqlite3.connect(engine.db_path)
    cursor = conn.cursor()

    # Determine if CN audit tables exist; if so, use them, else fall back to legacy
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='dhl_express_china_invoices'")
    china_invoices_exists = cursor.fetchone() is not None

    if china_invoices_exists:
        # Summary by invoice (use the per-invoice audit fields stored on each line; take MAX per invoice)
        cursor.execute('''
            SELECT audit_status,
                   COUNT(*) AS count,
                   AVG(CASE WHEN expected_cost_cny > 0 THEN variance_cny * 100.0 / expected_cost_cny END) AS avg_variance,
                   SUM(variance_cny) AS total_variance
            FROM (
                SELECT invoice_number,
                       MAX(audit_status) AS audit_status,
                       MAX(expected_cost_cny) AS expected_cost_cny,
                       MAX(variance_cny) AS variance_cny
                FROM dhl_express_china_invoices
                WHERE audit_status IS NOT NULL
                GROUP BY invoice_number
            ) inv
            GROUP BY audit_status
        ''')
        audit_summary = cursor.fetchall()

        # Recent audits (aggregate per invoice)
        cursor.execute('''
            SELECT inv.invoice_number,
                   MAX(inv.audit_timestamp) AS audit_timestamp,
                   MAX(inv.audit_status) AS audit_status,
                   SUM(inv.bcu_total) AS total_invoice_amount,
                   MAX(inv.variance_cny) AS total_variance,
                   CASE WHEN MAX(inv.expected_cost_cny) > 0 THEN MAX(inv.variance_cny) * 100.0 / MAX(inv.expected_cost_cny) ELSE 0 END AS variance_percentage
            FROM dhl_express_china_invoices inv
            WHERE inv.audit_status IS NOT NULL
            GROUP BY inv.invoice_number
            ORDER BY audit_timestamp DESC
            LIMIT 20
        ''')
        recent_audits = cursor.fetchall()

        # Convert to dictionaries
        summary_data = {}
        for row in audit_summary:
            status = row[0] or 'UNKNOWN'
            summary_data[status] = {
                'count': row[1] or 0,
                'avg_variance': round(row[2] or 0, 2),
                'total_variance': round(row[3] or 0, 2)
            }

        recent_audit_list = []
        for audit in recent_audits:
            variance_pct = float(audit[5] or 0)
            # Confidence: simple heuristic 1 - |variance%|/100, bounded [0,1]
            confidence = max(0.0, min(1.0, 1.0 - abs(variance_pct) / 100.0))
            recent_audit_list.append({
                'invoice_no': audit[0],
                'audit_timestamp': audit[1],
                'audit_status': audit[2],
                'total_invoice_amount': round(float(audit[3] or 0), 2),
                'total_variance': round(float(audit[4] or 0), 2),
                'variance_percentage': round(variance_pct, 2),
                'confidence_score': round(confidence, 2)
            })

        conn.close()
        return render_template('dhl_express_audit_summary.html',
                               summary=summary_data, recent_audits=recent_audit_list)
    else:
        # Legacy summary
        # Get audit results summary
        cursor.execute('''
            SELECT audit_status, COUNT(*) as count, 
                   AVG(variance_percentage) as avg_variance,
                   SUM(total_variance) as total_variance
            FROM dhl_express_audit_results
            GROUP BY audit_status
        ''')
        audit_summary = cursor.fetchall()
        
        # Get recent audits
        cursor.execute('''
            SELECT invoice_no, audit_timestamp, audit_status, 
                   total_invoice_amount, total_variance, variance_percentage,
                   confidence_score
            FROM dhl_express_audit_results
            ORDER BY audit_timestamp DESC
            LIMIT 20
        ''')
        recent_audits = cursor.fetchall()
        conn.close()
        
        # Convert to dictionaries
        summary_data = {}
        for summary in audit_summary:
            summary_data[summary[0]] = {
                'count': summary[1],
                'avg_variance': round(summary[2] or 0, 2),
                'total_variance': round(summary[3] or 0, 2)
            }
        
        recent_audit_list = []
        for audit in recent_audits:
            recent_audit_list.append({
                'invoice_no': audit[0],
                'audit_timestamp': audit[1],
                'audit_status': audit[2],
                'total_invoice_amount': round(audit[3], 2),
                'total_variance': round(audit[4], 2),
                'variance_percentage': round(audit[5], 2),
                'confidence_score': round(audit[6], 2)
            })
        
        return render_template('dhl_express_audit_summary.html',
                             summary=summary_data, recent_audits=recent_audit_list)


@dhl_express_routes.route('/dhl-express/run-invoice-audit', methods=['POST'])
@require_auth
def run_invoice_audit(user_data=None):
    """Re-audit all existing Chinese DHL Express invoices and save results."""
    try:
        # Ensure China engine is available
        if DHLExpressChinaAuditEngine is None:
            return jsonify({'success': False, 'error': 'China audit engine not available'}), 500

        engine_cn = DHLExpressChinaAuditEngine()
        conn = sqlite3.connect(engine_cn.db_path)
        cursor = conn.cursor()

        # Ensure CN invoice table exists
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='dhl_express_china_invoices'")
        if cursor.fetchone() is None:
            conn.close()
            return jsonify({'success': False, 'error': 'Chinese invoice table not found'}), 400

        # Get distinct invoice numbers
        cursor.execute('''
            SELECT DISTINCT invoice_number
            FROM dhl_express_china_invoices
            ORDER BY invoice_number
        ''')
        invoices = [row[0] for row in cursor.fetchall()]
        conn.close()

        completed = 0
        failed = []
        results = []
        for inv in invoices:
            try:
                audit = engine_cn.audit_invoice(inv)
                # Save results per-invoice
                engine_cn.save_audit_results(audit)
                results.append({
                    'invoice_number': inv,
                    'status': audit.get('status')
                })
                completed += 1
            except Exception as e:
                failed.append({'invoice_number': inv, 'error': str(e)})

        return jsonify({
            'success': True,
            'message': f'Completed audit for {completed} invoices',
            'completed': completed,
            'failed': failed,
            'results': results
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@dhl_express_routes.route('/dhl-express/invoice/<invoice_no>/download')
@require_auth
def download_invoice(invoice_no, user_data=None):
    """Download invoice details as Excel file"""
    import pandas as pd
    from io import BytesIO
    from flask import send_file
    
    engine = DHLExpressAuditEngine()
    
    # Get invoice details
    import sqlite3
    conn = sqlite3.connect(engine.db_path)
    cursor = conn.cursor()
    
    # Get invoice header
    cursor.execute('''
        SELECT DISTINCT invoice_no, invoice_date, company_name, account_number,
               shipper_details, receiver_details
        FROM dhl_express_invoices
        WHERE invoice_no = ?
    ''', (invoice_no,))
    
    header = cursor.fetchone()
    if not header:
        return "Invoice not found", 404
    
    # Get invoice lines
    cursor.execute('''
        SELECT line_number, dhl_product_description, amount, weight_charge,
               discount_amount, tax_amount, awb_number, weight, origin_code,
               destination_code, shipment_date, shipper_reference
        FROM dhl_express_invoices
        WHERE invoice_no = ?
        ORDER BY line_number
    ''', (invoice_no,))
    
    lines = cursor.fetchall()
    conn.close()
    
    # Create Excel file
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Invoice Header Sheet
        header_data = {
            'Field': ['Invoice Number', 'Invoice Date', 'Company Name', 'Account Number'],
            'Value': [header[0], header[1], header[2], header[3]]
        }
        header_df = pd.DataFrame(header_data)
        header_df.to_excel(writer, sheet_name='Invoice Header', index=False)
        
        # Invoice Lines Sheet
        lines_data = []
        for line in lines:
            lines_data.append({
                'Line Number': line[0],
                'Product Description': line[1],
                'Amount': line[2],
                'Weight Charge': line[3],
                'Discount Amount': line[4],
                'Tax Amount': line[5],
                'AWB Number': line[6],
                'Weight (kg)': line[7],
                'Origin': line[8],
                'Destination': line[9],
                'Shipment Date': line[10],
                'Shipper Reference': line[11]
            })
        
        lines_df = pd.DataFrame(lines_data)
        lines_df.to_excel(writer, sheet_name='Invoice Lines', index=False)
        
        # Shipper & Receiver Details
        def parse_address(address_string):
            if not address_string:
                return {}
            parts = [part.strip() for part in address_string.split(';') if part.strip()]
            return {
                'Company': parts[0] if len(parts) > 0 else '',
                'Address1': parts[1] if len(parts) > 1 else '',
                'Address2': parts[2] if len(parts) > 2 else '',
                'City': parts[3] if len(parts) > 3 else '',
                'State': parts[5] if len(parts) > 5 else '',
                'Postal Code': parts[6] if len(parts) > 6 else '',
                'Country': parts[7] if len(parts) > 7 else '',
                'Contact': parts[8] if len(parts) > 8 else '',
                'Email': parts[9] if len(parts) > 9 else ''
            }
        
        shipper = parse_address(header[4])
        receiver = parse_address(header[5])
        
        address_data = []
        for key in shipper.keys():
            address_data.append({
                'Field': key,
                'Shipper': shipper.get(key, ''),
                'Receiver': receiver.get(key, '')
            })
        
        address_df = pd.DataFrame(address_data)
        address_df.to_excel(writer, sheet_name='Addresses', index=False)
    
    output.seek(0)
    
    return send_file(
        output,
        as_attachment=True,
        download_name=f'DHL_Express_Invoice_{invoice_no}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@dhl_express_routes.route('/dhl-express/audit/<invoice_no>/download')
def download_audit_report(invoice_no, user_data=None):
    """Download audit report as Excel file"""
    import pandas as pd
    from io import BytesIO
    from flask import send_file
    import sqlite3
    import json
    
    engine = DHLExpressAuditEngine()
    
    # Get audit results from database
    conn = sqlite3.connect(engine.db_path)
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT * FROM dhl_express_audit_results 
        WHERE invoice_no = ?
        ORDER BY audit_timestamp DESC
        LIMIT 1
    ''', (invoice_no,))
    
    audit_row = cursor.fetchone()
    
    if not audit_row:
        # Try to run audit if no results found
        audit_result = engine.audit_invoice(invoice_no)
        if audit_result.get('status') == 'ERROR':
            return "Audit data not found", 404
        
        # Fetch the newly created audit result
        cursor.execute('''
            SELECT * FROM dhl_express_audit_results 
            WHERE invoice_no = ?
            ORDER BY audit_timestamp DESC
            LIMIT 1
        ''', (invoice_no,))
        
        audit_row = cursor.fetchone()
    
    if not audit_row:
        conn.close()
        return "Audit data not found", 404
    
    # Parse audit result data
    detailed_results = json.loads(audit_row[12]) if audit_row[12] else []
    
    # Create Excel file
    output = BytesIO()
    
    try:
        # Use xlsxwriter instead of openpyxl
        import xlsxwriter
        
        workbook = xlsxwriter.Workbook(output)
        
        # Audit Summary Sheet
        summary_sheet = workbook.add_worksheet('Audit Summary')
        
        # Add headers
        summary_sheet.write(0, 0, 'Field')
        summary_sheet.write(0, 1, 'Value')
        
        # Add data
        summary_data = [
            ['Invoice Number', str(audit_row[1])],  # invoice_no
            ['Audit Status', str(audit_row[8])],    # audit_status
            ['Total Invoice Amount', f"${float(audit_row[4]):,.2f}" if audit_row[4] is not None else "$0.00"],  # total_invoice_amount
            ['Total Expected Amount', f"${float(audit_row[5]):,.2f}" if audit_row[5] is not None else "$0.00"],  # total_expected_amount
            ['Total Variance', f"${float(audit_row[6]):,.2f}" if audit_row[6] is not None else "$0.00"],  # total_variance
            ['Variance Percentage', f"{float(audit_row[7]):,.2f}%" if audit_row[7] is not None else "0.00%"],  # variance_percentage
            ['Confidence Score', f"{float(audit_row[13]):,.1f}%" if audit_row[13] is not None else "0.0%"],  # confidence_score
            ['Lines Audited', str(audit_row[9]) if audit_row[9] is not None else "0"],   # line_items_audited
            ['Lines Passed', str(audit_row[10]) if audit_row[10] is not None else "0"],  # line_items_passed
            ['Lines Failed', str(audit_row[11]) if audit_row[11] is not None else "0"],  # line_items_failed
            ['Audit Date', str(audit_row[3]) if audit_row[3] is not None else ""]        # audit_timestamp
        ]
        
        for row, (field, value) in enumerate(summary_data, start=1):
            summary_sheet.write(row, 0, field)
            summary_sheet.write(row, 1, value)
        
        # Detailed Results Sheet
        detail_sheet = workbook.add_worksheet('Line-by-Line Results')
        
        if detailed_results:
            # Headers
            headers = ['Line Number', 'Product Description', 'Invoiced Amount', 
                      'Expected Amount', 'Variance', 'Audit Result', 'Comments']
            
            for col, header in enumerate(headers):
                detail_sheet.write(0, col, header)
            
            # Data
            for row, line_result in enumerate(detailed_results, start=1):
                detail_sheet.write(row, 0, str(line_result.get('line_number', '')))
                detail_sheet.write(row, 1, str(line_result.get('description', '')))
                invoiced = line_result.get('invoiced', 0)
                detail_sheet.write(row, 2, f"${float(invoiced):,.2f}" if invoiced is not None else "$0.00")
                expected = line_result.get('expected', 0)
                detail_sheet.write(row, 3, f"${float(expected):,.2f}" if expected is not None else "$0.00")
                variance = line_result.get('variance', 0)
                detail_sheet.write(row, 4, f"${float(variance):,.2f}" if variance is not None else "$0.00")
                detail_sheet.write(row, 5, str(line_result.get('result', '')))
                comments = line_result.get('comments', [])
                comment_text = '; '.join(comments) if isinstance(comments, list) else str(comments)
                detail_sheet.write(row, 6, comment_text)
        else:
            detail_sheet.write(0, 0, 'No detailed line items found')
        
        workbook.close()
        
    except Exception as e:
        conn.close()
        return f"Error creating Excel file: {str(e)}", 500
    
    conn.close()
    output.seek(0)
    
    return send_file(
        output,
        as_attachment=True,
        download_name=f'DHL_Express_Audit_Report_{invoice_no}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# Batch Audit Routes
@dhl_express_routes.route('/dhl-express/batch-audit')
@require_auth
def batch_audit_dashboard(user_data=None):
    """Display batch audit dashboard"""
    engine = DHLExpressAuditEngine()
    
    # Get audit status summary
    audit_summary = engine.get_audit_status_summary()
    unaudited_invoices = engine.get_unaudited_invoices()
    
    return render_template('dhl_express_batch_audit.html', 
                         summary=audit_summary,
                         unaudited_count=len(unaudited_invoices),
                         unaudited_invoices=unaudited_invoices[:10])  # Show first 10

@dhl_express_routes.route('/dhl-express/batch-audit/status')
def get_batch_audit_status():
    """Get current batch audit status (API endpoint)"""
    engine = DHLExpressAuditEngine()
    summary = engine.get_audit_status_summary()
    unaudited_invoices = engine.get_unaudited_invoices()
    
    return jsonify({
        'success': True,
        'summary': summary,
        'unaudited_invoices': len(unaudited_invoices),
        'sample_unaudited': unaudited_invoices[:5]
    })

@dhl_express_routes.route('/dhl-express/batch-audit/run', methods=['POST'])
def run_batch_audit():
    """Run batch audit on all unaudited invoices"""
    try:
        engine = DHLExpressAuditEngine()
        
        # Get request parameters
        data = request.get_json() or {}
        audit_type = data.get('audit_type', 'all_unaudited')  # 'all_unaudited' or 'specific_invoices'
        specific_invoices = data.get('invoice_list', [])
        
        if audit_type == 'specific_invoices' and specific_invoices:
            # Audit specific invoices
            result = engine.audit_batch(specific_invoices)
        else:
            # Audit all unaudited invoices
            result = engine.audit_all_unaudited_invoices()
        
        return jsonify({
            'success': True,
            'batch_result': result
        })
        
    except Exception as e:
        import traceback
        print(f"Batch audit error: {traceback.format_exc()}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@dhl_express_routes.route('/dhl-express/batch-audit/rerun', methods=['POST'])
@require_auth
def rerun_batch_audit(user_data=None):
    """Re-run batch audit on all invoices to refresh cached results"""
    try:
        engine = DHLExpressAuditEngine()
        
        # Clear all existing audit results
        conn = sqlite3.connect(engine.db_path)
        cursor = conn.cursor()
        
        # Get all unique invoice numbers
        cursor.execute('SELECT DISTINCT invoice_no FROM dhl_express_invoices')
        invoice_numbers = [row[0] for row in cursor.fetchall()]
        
        # Clear existing audit results
        cursor.execute('DELETE FROM dhl_express_audit_results')
        conn.commit()
        conn.close()
        
        # Re-audit all invoices
        result = engine.audit_batch(invoice_numbers)
        
        return jsonify({
            'success': True,
            'message': f'Successfully re-audited {result["completed"]} invoices',
            'batch_result': result
        })
        
    except Exception as e:
        import traceback
        print(f"Re-run batch audit error: {traceback.format_exc()}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@dhl_express_routes.route('/dhl-express/batch-audit/results')
@require_auth
def batch_audit_results(user_data=None):
    """Display batch audit results"""
    engine = DHLExpressAuditEngine()
    
    # Get all audit results with pagination
    page = request.args.get('page', 1, type=int)
    per_page = 50
    offset = (page - 1) * per_page
    
    conn = sqlite3.connect(engine.db_path)
    cursor = conn.cursor()
    
    # Get total count
    cursor.execute('SELECT COUNT(*) FROM dhl_express_audit_results')
    total_audits = cursor.fetchone()[0]
    
    # Get paginated results
    cursor.execute('''
        SELECT invoice_no, audit_timestamp, total_invoice_amount, 
               total_expected_amount, total_variance, variance_percentage,
               audit_status, line_items_audited, confidence_score
        FROM dhl_express_audit_results
        ORDER BY audit_timestamp DESC
        LIMIT ? OFFSET ?
    ''', (per_page, offset))
    
    audit_results = []
    for row in cursor.fetchall():
        audit_results.append({
            'invoice_no': row[0],
            'audit_timestamp': row[1],
            'total_invoice_amount': row[2],
            'total_expected_amount': row[3],
            'total_variance': row[4],
            'variance_percentage': row[5],
            'audit_status': row[6],
            'line_items_audited': row[7],
            'confidence_score': row[8]
        })
    
    conn.close()
    
    # Calculate pagination info
    total_pages = (total_audits + per_page - 1) // per_page
    has_prev = page > 1
    has_next = page < total_pages
    
    return render_template('dhl_express_batch_results.html',
                         audit_results=audit_results,
                         page=page,
                         total_pages=total_pages,
                         has_prev=has_prev,
                         has_next=has_next,
                         total_audits=total_audits)

@dhl_express_routes.route('/dhl-express/batch-audit/export')
@require_auth
def export_batch_audit_results(user_data=None):
    """Export all batch audit results to Excel"""
    try:
        engine = DHLExpressAuditEngine()
        conn = sqlite3.connect(engine.db_path)
        
        # Get all audit results
        query = '''
            SELECT invoice_no, audit_timestamp, total_invoice_amount,
                   total_expected_amount, total_variance, variance_percentage,
                   audit_status, line_items_audited, line_items_passed,
                   line_items_failed, confidence_score
            FROM dhl_express_audit_results
            ORDER BY audit_timestamp DESC
        '''
        
        df = pd.read_sql_query(query, conn)
        conn.close()
        
        # Create Excel file
        output = BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df.to_excel(writer, sheet_name='Batch Audit Results', index=False)
            
            # Get the workbook and worksheet objects
            workbook = writer.book
            worksheet = writer.sheets['Batch Audit Results']
            
            # Add formatting
            header_format = workbook.add_format({
                'bold': True,
                'text_wrap': True,
                'valign': 'top',
                'fg_color': '#D7E4BC',
                'border': 1
            })
            
            # Write headers with formatting
            for col_num, value in enumerate(df.columns.values):
                worksheet.write(0, col_num, value, header_format)
                
            # Auto-adjust column widths
            for i, col in enumerate(df.columns):
                column_len = max(df[col].astype(str).str.len().max(), len(col)) + 2
                worksheet.set_column(i, i, min(column_len, 50))
        
        output.seek(0)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return send_file(
            output,
            as_attachment=True,
            download_name=f'DHL_Express_Batch_Audit_Results_{timestamp}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'Error exporting batch audit results: {str(e)}', 'error')
        return redirect(url_for('dhl_express.batch_audit_results'))

@dhl_express_routes.route('/dhl-express/batch-audit/export-detailed')
def export_detailed_batch_audit_results(user_data=None):
    """Export detailed line-by-line batch audit results for finance review"""
    try:
        import xlsxwriter
        from io import BytesIO
        
        engine = DHLExpressAuditEngine()
        conn = sqlite3.connect(engine.db_path)
        cursor = conn.cursor()
        
        # Get all audit results with detailed breakdown
        cursor.execute('''
            SELECT r.invoice_no, r.audit_timestamp, r.total_invoice_amount,
                   r.total_expected_amount, r.total_variance, r.variance_percentage,
                   r.audit_status, r.line_items_audited, r.line_items_passed,
                   r.line_items_failed, r.confidence_score, r.detailed_results,
                   i.company_name, i.account_number
            FROM dhl_express_audit_results r
            LEFT JOIN (
                SELECT DISTINCT invoice_no, company_name, account_number 
                FROM dhl_express_invoices
            ) i ON r.invoice_no = i.invoice_no
            ORDER BY r.audit_timestamp DESC
        ''')
        
        audit_results = cursor.fetchall()
        
        # Create Excel file with detailed breakdown
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output)
        
        # Define formats
        header_format = workbook.add_format({
            'bold': True,
            'text_wrap': True,
            'valign': 'top',
            'fg_color': '#366092',
            'font_color': 'white',
            'border': 1,
            'font_size': 11
        })
        
        summary_header_format = workbook.add_format({
            'bold': True,
            'text_wrap': True,
            'valign': 'top',
            'fg_color': '#70AD47',
            'font_color': 'white',
            'border': 1,
            'font_size': 11
        })
        
        currency_format = workbook.add_format({
            'num_format': '$#,##0.00',
            'border': 1
        })
        
        percentage_format = workbook.add_format({
            'num_format': '0.00%',
            'border': 1
        })
        
        pass_format = workbook.add_format({
            'bg_color': '#C6EFCE',
            'border': 1
        })
        
        fail_format = workbook.add_format({
            'bg_color': '#FFC7CE',
            'border': 1
        })
        
        review_format = workbook.add_format({
            'bg_color': '#FFEB9C',
            'border': 1
        })
        
        # Summary Sheet
        summary_sheet = workbook.add_worksheet('Audit Summary')
        
        summary_headers = [
            'Invoice Number', 'Company Name', 'Account Number', 'Audit Date',
            'Total Invoice Amount', 'Total Expected Amount', 'Total Variance',
            'Variance %', 'Status', 'Lines Audited', 'Lines Passed', 
            'Lines Failed', 'Confidence Score', 'Finance Action Required'
        ]
        
        # Write summary headers
        for col, header in enumerate(summary_headers):
            summary_sheet.write(0, col, header, summary_header_format)
        
        # Write summary data
        row = 1
        for audit_result in audit_results:
            invoice_no = audit_result[0]
            audit_timestamp = audit_result[1]
            total_invoice_amount = float(audit_result[2]) if audit_result[2] else 0
            total_expected_amount = float(audit_result[3]) if audit_result[3] else 0
            total_variance = float(audit_result[4]) if audit_result[4] else 0
            variance_percentage = float(audit_result[5]) if audit_result[5] else 0
            audit_status = audit_result[6]
            line_items_audited = audit_result[7]
            line_items_passed = audit_result[8]
            line_items_failed = audit_result[9]
            confidence_score = float(audit_result[10]) if audit_result[10] else 0
            company_name = audit_result[12] or 'Unknown'
            account_number = audit_result[13] or 'Unknown'
            
            # Determine finance action
            if audit_status == 'FAIL' or total_variance < -50:  # DHL overcharged significantly
                finance_action = 'DISPUTE REQUIRED - Request Credit'
            elif total_variance < -10:  # DHL overcharged moderately
                finance_action = 'REVIEW REQUIRED - Consider Dispute'
            elif total_variance > 50:  # DHL undercharged significantly
                finance_action = 'VERIFY ACCURACY - Possible Underbilling'
            else:
                finance_action = 'APPROVE FOR PAYMENT'
            
            # Choose status format
            status_format = pass_format if audit_status == 'PASS' else (
                fail_format if audit_status == 'FAIL' else review_format
            )
            
            summary_sheet.write(row, 0, invoice_no)
            summary_sheet.write(row, 1, company_name)
            summary_sheet.write(row, 2, account_number)
            summary_sheet.write(row, 3, audit_timestamp)
            summary_sheet.write(row, 4, total_invoice_amount, currency_format)
            summary_sheet.write(row, 5, total_expected_amount, currency_format)
            summary_sheet.write(row, 6, total_variance, currency_format)
            summary_sheet.write(row, 7, variance_percentage / 100, percentage_format)
            summary_sheet.write(row, 8, audit_status, status_format)
            summary_sheet.write(row, 9, line_items_audited)
            summary_sheet.write(row, 10, line_items_passed)
            summary_sheet.write(row, 11, line_items_failed)
            summary_sheet.write(row, 12, confidence_score)
            summary_sheet.write(row, 13, finance_action)
            
            row += 1
        
        # Set column widths for summary
        summary_sheet.set_column('A:A', 18)  # Invoice Number
        summary_sheet.set_column('B:B', 25)  # Company Name
        summary_sheet.set_column('C:C', 15)  # Account Number
        summary_sheet.set_column('D:D', 20)  # Audit Date
        summary_sheet.set_column('E:G', 18)  # Amounts
        summary_sheet.set_column('H:H', 12)  # Variance %
        summary_sheet.set_column('I:I', 10)  # Status
        summary_sheet.set_column('J:L', 12)  # Line counts
        summary_sheet.set_column('M:M', 15)  # Confidence
        summary_sheet.set_column('N:N', 30)  # Finance Action
        
        # Detailed Line Items Sheet
        detail_sheet = workbook.add_worksheet('Line-by-Line Details')
        
        detail_headers = [
            'Invoice Number', 'Line Number', 'Product Description', 'AWB Number',
            'Origin', 'Destination', 'Weight (kg)', 'Invoiced Amount', 
            'Expected Amount', 'Variance', 'Status', 'Comments', 'Finance Notes'
        ]
        
        # Write detail headers
        for col, header in enumerate(detail_headers):
            detail_sheet.write(0, col, header, header_format)
        
        # Write detailed line items
        detail_row = 1
        for audit_result in audit_results:
            invoice_no = audit_result[0]
            detailed_results = json.loads(audit_result[11]) if audit_result[11] else []
            
            # Get additional invoice details
            cursor.execute('''
                SELECT awb_number, origin_code, destination_code, weight
                FROM dhl_express_invoices
                WHERE invoice_no = ?
                ORDER BY line_number
            ''', (invoice_no,))
            invoice_details = cursor.fetchall()
            invoice_detail_dict = {f"{inv[0]}": inv for inv in invoice_details}
            
            for line_result in detailed_results:
                line_number = line_result.get('line_number', '')
                description = line_result.get('description', '')
                invoiced = float(line_result.get('invoiced', 0))
                expected = float(line_result.get('expected', 0))
                variance = float(line_result.get('variance', 0))
                status = line_result.get('result', '')
                comments = line_result.get('comments', [])
                
                # Format comments
                if isinstance(comments, list):
                    comment_text = '; '.join(comments)
                else:
                    comment_text = str(comments)
                
                # Get additional details from invoice
                awb_number = ''
                origin = ''
                destination = ''
                weight = 0
                
                # Try to match by AWB or description
                for inv_detail in invoice_details:
                    if inv_detail[0]:  # Has AWB
                        awb_number = inv_detail[0]
                        origin = inv_detail[1] or ''
                        destination = inv_detail[2] or ''
                        weight = inv_detail[3] or 0
                        break
                
                # Finance notes based on variance and status
                if status == 'FAIL' and variance < -10:
                    finance_notes = 'DISPUTE - DHL Overcharged'
                elif variance < -5:
                    finance_notes = 'Review for potential dispute'
                elif variance > 10:
                    finance_notes = 'Verify - Possible underbilling'
                else:
                    finance_notes = 'Approved'
                
                # Choose format based on status
                line_format = pass_format if status == 'PASS' else (
                    fail_format if status == 'FAIL' else review_format
                )
                
                detail_sheet.write(detail_row, 0, invoice_no)
                detail_sheet.write(detail_row, 1, str(line_number))
                detail_sheet.write(detail_row, 2, description)
                detail_sheet.write(detail_row, 3, awb_number)
                detail_sheet.write(detail_row, 4, origin)
                detail_sheet.write(detail_row, 5, destination)
                detail_sheet.write(detail_row, 6, weight)
                detail_sheet.write(detail_row, 7, invoiced, currency_format)
                detail_sheet.write(detail_row, 8, expected, currency_format)
                detail_sheet.write(detail_row, 9, variance, currency_format)
                detail_sheet.write(detail_row, 10, status, line_format)
                detail_sheet.write(detail_row, 11, comment_text)
                detail_sheet.write(detail_row, 12, finance_notes)
                
                detail_row += 1
        
        # Set column widths for details
        detail_sheet.set_column('A:A', 18)  # Invoice Number
        detail_sheet.set_column('B:B', 10)  # Line Number
        detail_sheet.set_column('C:C', 35)  # Product Description
        detail_sheet.set_column('D:D', 15)  # AWB Number
        detail_sheet.set_column('E:F', 10)  # Origin/Destination
        detail_sheet.set_column('G:G', 12)  # Weight
        detail_sheet.set_column('H:J', 15)  # Amounts
        detail_sheet.set_column('K:K', 10)  # Status
        detail_sheet.set_column('L:L', 40)  # Comments
        detail_sheet.set_column('M:M', 25)  # Finance Notes
        
        # Disputes Summary Sheet (for finance team focus)
        disputes_sheet = workbook.add_worksheet('Disputes Required')
        
        dispute_headers = [
            'Priority', 'Invoice Number', 'Company Name', 'Total Variance',
            'Description', 'Recommended Action', 'Amount to Dispute'
        ]
        
        # Write dispute headers
        for col, header in enumerate(dispute_headers):
            disputes_sheet.write(0, col, header, summary_header_format)
        
        # Collect disputes
        dispute_row = 1
        for audit_result in audit_results:
            total_variance = float(audit_result[4]) if audit_result[4] else 0
            if total_variance < -5:  # DHL overcharged by more than $5
                invoice_no = audit_result[0]
                company_name = audit_result[12] or 'Unknown'
                audit_status = audit_result[6]
                
                if total_variance < -50:
                    priority = 'HIGH'
                    action = 'IMMEDIATE DISPUTE REQUIRED'
                elif total_variance < -20:
                    priority = 'MEDIUM'
                    action = 'DISPUTE RECOMMENDED'
                else:
                    priority = 'LOW'
                    action = 'REVIEW FOR DISPUTE'
                
                disputes_sheet.write(dispute_row, 0, priority)
                disputes_sheet.write(dispute_row, 1, invoice_no)
                disputes_sheet.write(dispute_row, 2, company_name)
                disputes_sheet.write(dispute_row, 3, total_variance, currency_format)
                disputes_sheet.write(dispute_row, 4, f'DHL overcharged by ${abs(total_variance):,.2f}')
                disputes_sheet.write(dispute_row, 5, action)
                disputes_sheet.write(dispute_row, 6, abs(total_variance), currency_format)
                
                dispute_row += 1
        
        # Set column widths for disputes
        disputes_sheet.set_column('A:A', 10)  # Priority
        disputes_sheet.set_column('B:B', 18)  # Invoice Number
        disputes_sheet.set_column('C:C', 25)  # Company Name
        disputes_sheet.set_column('D:D', 15)  # Total Variance
        disputes_sheet.set_column('E:E', 30)  # Description
        disputes_sheet.set_column('F:F', 25)  # Action
        disputes_sheet.set_column('G:G', 18)  # Amount to Dispute
        
        workbook.close()
        conn.close()
        
        output.seek(0)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return send_file(
            output,
            as_attachment=True,
            download_name=f'DHL_Express_Detailed_Audit_Results_For_Finance_{timestamp}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return f"Error creating detailed audit report: {str(e)}", 500
