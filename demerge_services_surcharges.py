#!/usr/bin/env python3
"""
Demerge and populate Services & Surcharges Excel sheet to make it computer-friendly
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
import os

def find_services_sheet(file_path):
    """Find the Services & Surcharges sheet by checking similar names"""
    try:
        wb = load_workbook(file_path, data_only=True)
        sheet_names = wb.sheetnames
        
        # Look for sheets that might contain services/surcharges
        possible_names = [
            's&s published', 'services published', 'surcharges published',
            's&s', 'services', 'surcharges', 'premium services',
            'service charges', 'additional services'
        ]
        
        for sheet_name in sheet_names:
            sheet_lower = sheet_name.lower()
            for possible in possible_names:
                if possible in sheet_lower:
                    print(f"Found potential services sheet: '{sheet_name}'")
                    return sheet_name
        
        print("Available sheets:")
        for i, name in enumerate(sheet_names):
            print(f"  {i+1}: {name}")
        
        return None
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return None

def demerge_services_sheet(file_path, sheet_name, output_file=None):
    """
    Demerge cells and populate blank rows in Services & Surcharges sheet
    """
    try:
        print(f"Reading sheet '{sheet_name}' from {file_path}")
        
        # Read the sheet with all data
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        
        print(f"Original sheet dimensions: {df.shape}")
        print("\nFirst 10 rows of original data:")
        for i in range(min(10, len(df))):
            print(f"Row {i}: {df.iloc[i].tolist()[:6]}...")  # Show first 6 columns
        
        # Find the header row (look for common service charge headers)
        header_row = None
        for i in range(min(20, len(df))):
            row_str = ' '.join([str(cell).upper() for cell in df.iloc[i] if pd.notna(cell)])
            # Look for row with multiple standard column headers
            if 'CODE' in row_str and 'NAME' in row_str and any(word in row_str for word in ['DESCRIPTION', 'CHARGE', 'PRICE', 'RATE']):
                header_row = i
                print(f"\nFound header row at index {i}: {df.iloc[i].tolist()}")
                break
        
        if header_row is None:
            print("Could not find proper header row. Please check the sheet structure.")
            return None, None
        
        # Extract headers and data
        headers = df.iloc[header_row].tolist()
        data_df = df.iloc[header_row + 1:].copy()
        data_df.columns = range(len(data_df.columns))
        
        print(f"\nHeaders: {headers}")
        print(f"Data shape after header extraction: {data_df.shape}")
        
        # Identify key columns (service code, description, rate info)
        service_code_col = None
        description_col = None
        rate_cols = []
        
        for i, header in enumerate(headers):
            if pd.notna(header):
                header_str = str(header).upper()
                if 'CODE' in header_str and service_code_col is None:
                    service_code_col = i
                elif any(word in header_str for word in ['DESCRIPTION', 'SERVICE', 'NAME']) and description_col is None:
                    description_col = i
                elif any(word in header_str for word in ['RATE', 'CHARGE', 'AMOUNT', 'DOMESTIC', 'INTERNATIONAL']):
                    rate_cols.append(i)
        
        print(f"\nIdentified columns:")
        print(f"  Service code column: {service_code_col}")
        print(f"  Description column: {description_col}")
        print(f"  Rate columns: {rate_cols}")
        
        # Create the demerged dataframe
        demerged_data = []
        current_service_code = None
        current_description = None
        
        for index, row in data_df.iterrows():
            # Check if this row has a service code
            if service_code_col is not None and pd.notna(row.iloc[service_code_col]):
                current_service_code = row.iloc[service_code_col]
            
            # Check if this row has a description
            if description_col is not None and pd.notna(row.iloc[description_col]):
                current_description = row.iloc[description_col]
            
            # Check if this row has any rate data
            has_rate_data = False
            if rate_cols:
                has_rate_data = any(pd.notna(row.iloc[col]) and str(row.iloc[col]).strip() != '' 
                                  for col in rate_cols)
            else:
                # If no rate columns identified, check if any cell in the row has data
                has_rate_data = any(pd.notna(cell) and str(cell).strip() != '' for cell in row)
            
            # If this row has rate data, create a complete record
            if has_rate_data:
                new_row = row.copy()
                
                # Fill in the service code and description if they're missing
                if service_code_col is not None:
                    if pd.isna(new_row.iloc[service_code_col]) or str(new_row.iloc[service_code_col]).strip() == '':
                        new_row.iloc[service_code_col] = current_service_code
                
                if description_col is not None:
                    if pd.isna(new_row.iloc[description_col]) or str(new_row.iloc[description_col]).strip() == '':
                        new_row.iloc[description_col] = current_description
                
                demerged_data.append(new_row.tolist())
        
        # Create the final dataframe
        if demerged_data:
            demerged_df = pd.DataFrame(demerged_data, columns=headers)
            
            print(f"\nDemerged data shape: {demerged_df.shape}")
            print("\nFirst 10 rows of demerged data:")
            print(demerged_df.head(10).to_string())
            
            # Save to new file
            if output_file is None:
                base_name = os.path.splitext(file_path)[0]
                output_file = f"{base_name}_demerged.xlsx"
            
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                demerged_df.to_excel(writer, sheet_name='Services_Demerged', index=False)
            
            print(f"\nDemerged data saved to: {output_file}")
            
            # Show summary statistics
            if service_code_col is not None:
                unique_codes = demerged_df.iloc[:, service_code_col].dropna().unique()
                print(f"\nFound {len(unique_codes)} unique service codes:")
                for code in sorted(unique_codes):
                    if pd.notna(code):
                        print(f"  {code}")
            
            return output_file, demerged_df
        else:
            print("No data rows found with rate information.")
            return None, None
            
    except Exception as e:
        print(f"Error processing sheet: {e}")
        import traceback
        traceback.print_exc()
        return None, None

def main():
    """Main function to process the services sheet"""
    
    # File path - adjust as needed
    file_path = 'uploads/ID_104249_01_AP_V01_Commscope_AU_20241113-074659-898.xlsx'
    
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        print("Please check the file path.")
        return
    
    # Find the services sheet
    sheet_name = find_services_sheet(file_path)
    
    if sheet_name is None:
        print("\nCould not automatically find Services & Surcharges sheet.")
        print("Please specify the exact sheet name manually.")
        return
    
    # Process the sheet
    output_file, demerged_df = demerge_services_sheet(file_path, sheet_name)
    
    if output_file:
        print(f"\n✅ Successfully created demerged file: {output_file}")
        print("\nThis file should now be computer-friendly with:")
        print("- No merged cells")
        print("- Complete service information in each row") 
        print("- Proper service code and description associations")
        print("\nYou can now use this file for automated processing!")
    else:
        print("\n❌ Failed to create demerged file.")

if __name__ == "__main__":
    main()
